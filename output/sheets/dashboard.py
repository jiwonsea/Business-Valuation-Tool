"""Sheet 7: Dashboard — executive summary, charts, football field."""

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ._ctx import Ctx
from ..excel_styles import (
    NAVY, BLUE_FILL, GREEN_FILL, RED_FILL, GRAY_FILL,
    SECTION_FONT, NOTE_FONT, WHITE_FONT,
    NUM_FMT,
    style_header_row, write_cell,
)

CHART_ROWS = 18
CHART_GAP = 3


def sheet_dashboard(ctx: Ctx):
    ws = ctx.wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = NAVY
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    method_labels = {
        "sotp": "SOTP (Sum-of-the-Parts)",
        "dcf_primary": "DCF (Discounted Cash Flow)",
        "ddm": "DDM (배당할인모델)",
        "rim": "RIM (잔여이익모델)",
        "nav": "NAV (순자산가치)",
        "multiples": "Multiples (상대가치평가)",
    }
    method_desc = method_labels.get(ctx.method, ctx.method.upper())
    if ctx.method == "sotp" and ctx.result.dcf is not None:
        method_desc = "SOTP + DCF (Cross-Validation)"

    write_cell(ws, 1, 1, f"{ctx.vi.company.name} 기업가치평가 Dashboard",
               font=Font(bold=True, size=16, color=NAVY))
    write_cell(ws, 2, 1, f"분석일: {ctx.vi.company.analysis_date}  |  {method_desc}",
               font=NOTE_FONT)

    r = 4

    # ── Key conclusion ──
    primary_value, primary_label = _get_primary_value(ctx)
    write_cell(ws, r, 1, primary_label,
               font=Font(bold=True, size=14, color=NAVY))
    write_cell(ws, r, 2, f"{primary_value:,}{ctx.currency_sym}",
               font=Font(bold=True, size=18, color="27AE60"), fill=GREEN_FILL)

    # ── Scenario summary (only when available) ──
    sc_header_row = None
    if ctx.sc_codes and ctx.result.scenarios:
        r += 2
        write_cell(ws, r, 1, "시나리오별 주당 가치", font=SECTION_FONT); r += 1
        sc_sum_headers = ["시나리오", "주당가치", "확률", "가중기여"]
        for c, h in enumerate(sc_sum_headers, 1):
            write_cell(ws, r, c, h)
        style_header_row(ws, r, 4)
        sc_header_row = r

        for sc_code in ctx.sc_codes:
            r += 1
            sc = ctx.vi.scenarios[sc_code]
            sr = ctx.result.scenarios[sc_code]
            write_cell(ws, r, 1, f"{sc_code}: {sc.name}")
            write_cell(ws, r, 2, sr.post_dlom, fmt=NUM_FMT,
                       fill=GREEN_FILL if sr.post_dlom > 0 else RED_FILL)
            write_cell(ws, r, 3, f"{sc.prob}%")
            write_cell(ws, r, 4, sr.weighted, fmt=NUM_FMT)

        r += 1
        write_cell(ws, r, 1, "합계", bold=True)
        write_cell(ws, r, 2, ctx.result.weighted_value, fmt=NUM_FMT, bold=True, fill=GREEN_FILL)
        write_cell(ws, r, 3, "100%", bold=True)
        write_cell(ws, r, 4, ctx.result.weighted_value, fmt=NUM_FMT, bold=True)

    # ── Valuation summary by method ──
    r += 2
    ev_data_start = None
    ev_data_end = None

    if ctx.method == "ddm" and ctx.result.ddm:
        ddm = ctx.result.ddm
        write_cell(ws, r, 1, "DDM 밸류에이션 요약", font=SECTION_FONT); r += 1
        items = [("주당 배당금 (DPS)", f"{ddm.dps:,.0f}")]
        if ddm.buyback_per_share > 0:
            items.append(("자사주매입/주", f"{ddm.buyback_per_share:,.0f}"))
            items.append(("Total Payout/주", f"{ddm.total_payout:,.0f}"))
        items += [
            ("배당 성장률", f"{ddm.growth:.2f}%"),
            ("자기자본비용 (Ke)", f"{ddm.ke:.2f}%"),
            ("주당 내재가치 (DDM)", f"{ddm.equity_per_share:,}"),
        ]
        for label, val in items:
            write_cell(ws, r, 1, label)
            is_result = "내재가치" in label
            write_cell(ws, r, 2, val, fill=GREEN_FILL if is_result else BLUE_FILL, bold=is_result)
            r += 1

    elif ctx.method == "rim" and ctx.result.rim:
        rim = ctx.result.rim
        write_cell(ws, r, 1, "RIM 밸류에이션 요약", font=SECTION_FONT); r += 1
        items = [
            ("장부가치 (BV₀)", f"{rim.bv_current:,}"),
            ("자기자본비용 (Ke)", f"{rim.ke:.2f}%"),
            ("PV(잔여이익)", f"{rim.pv_ri_sum:,}"),
            ("PV(Terminal)", f"{rim.pv_terminal:,}"),
            ("자기자본가치", f"{rim.equity_value:,}"),
            ("주당 내재가치 (RIM)", f"{rim.per_share:,}"),
        ]
        for label, val in items:
            write_cell(ws, r, 1, label)
            is_result = "내재가치" in label
            write_cell(ws, r, 2, val, fill=GREEN_FILL if is_result else BLUE_FILL, bold=is_result)
            r += 1

    elif ctx.method == "nav" and ctx.result.nav:
        nav = ctx.result.nav
        write_cell(ws, r, 1, "NAV 밸류에이션 요약", font=SECTION_FONT); r += 1
        items = [
            ("총자산 (장부)", f"{nav.total_assets:,}"),
            ("(+) 재평가 조정", f"{nav.revaluation:,}"),
            ("(-) 총부채", f"{nav.total_liabilities:,}"),
            ("순자산가치 (NAV)", f"{nav.nav:,}"),
            ("주당 NAV", f"{nav.per_share:,}"),
        ]
        for label, val in items:
            write_cell(ws, r, 1, label)
            is_result = "주당" in label
            write_cell(ws, r, 2, val, fill=GREEN_FILL if is_result else BLUE_FILL, bold=is_result)
            r += 1

    elif ctx.method == "multiples" and ctx.result.multiples_primary:
        mp = ctx.result.multiples_primary
        write_cell(ws, r, 1, "Multiples 밸류에이션 요약", font=SECTION_FONT); r += 1
        items = [
            ("방법론", mp.primary_multiple_method),
            ("적용 멀티플", f"{mp.multiple:.1f}x"),
            ("Equity Value", f"{mp.equity_value:,}"),
            ("주당 가치", f"{mp.per_share:,}"),
        ]
        for label, val in items:
            write_cell(ws, r, 1, label)
            is_result = "주당" in label
            write_cell(ws, r, 2, val, fill=GREEN_FILL if is_result else BLUE_FILL, bold=is_result)
            r += 1

    elif ctx.method == "dcf_primary" and ctx.result.dcf:
        dcf = ctx.result.dcf
        write_cell(ws, r, 1, "DCF 밸류에이션 요약", font=SECTION_FONT); r += 1
        items = [
            ("PV(FCFF)", f"{dcf.pv_fcff_sum:,}"),
            ("PV(Terminal)", f"{dcf.pv_terminal:,}"),
            ("DCF EV", f"{dcf.ev_dcf:,}"),
            ("WACC", f"{dcf.wacc:.2f}%"),
            ("Terminal Growth", f"{dcf.terminal_growth:.1f}%"),
        ]
        for label, val in items:
            write_cell(ws, r, 1, label)
            is_ev = label == "DCF EV"
            write_cell(ws, r, 2, val, fill=GREEN_FILL if is_ev else BLUE_FILL, bold=is_ev)
            r += 1

    else:
        # SOTP (default) -- Mixed Method handling
        has_mixed_dash = ctx.result.sotp and any(
            getattr(s, "method", "ev_ebitda") != "ev_ebitda"
            for s in ctx.result.sotp.values()
        )
        section_title = "SOTP 밸류에이션 구성 (Mixed)" if has_mixed_dash else f"Enterprise Value 구성 ({ctx.unit})"
        write_cell(ws, r, 1, section_title, font=SECTION_FONT); r += 1
        ev_data_start = r
        active_segs = []
        if ctx.result.sotp:
            active_segs = [c for c in ctx.seg_codes if ctx.result.sotp.get(c) and ctx.result.sotp[c].ev > 0]
            for code in active_segs:
                s = ctx.result.sotp[code]
                method = getattr(s, "method", "ev_ebitda")
                m_label = {"ev_ebitda": "EV/EBITDA", "pbv": "P/BV", "pe": "P/E", "ev_revenue": "EV/Revenue"}.get(method, "")
                write_cell(ws, r, 1, f"{ctx.seg_names[code]} ({m_label} {s.multiple:.1f}x)")
                write_cell(ws, r, 2, s.ev, fmt=NUM_FMT)
                r += 1
        ev_data_end = r - 1
        write_cell(ws, r, 1, "Total EV", bold=True)
        write_cell(ws, r, 2, ctx.result.total_ev, fmt=NUM_FMT, bold=True, fill=GREEN_FILL)

    # ── Key financial metrics ──
    r += 2
    cons_by = ctx.cons[ctx.by]
    total_da = cons_by["dep"] + cons_by["amort"]
    ebitda = cons_by["op"] + total_da
    write_cell(ws, r, 1, f"핵심 재무지표 ({ctx.by})", font=SECTION_FONT); r += 1

    # Mixed SOTP: display effective net debt
    _is_mixed_dash = bool(ctx.vi.segment_net_debt) and any(
        info.get("method") in ("pbv", "pe") for info in ctx.vi.segments.values()
    )
    kpis = [
        ("매출액", cons_by["revenue"]),
        ("영업이익", cons_by["op"]),
        ("EBITDA", ebitda),
    ]
    if _is_mixed_dash:
        fin_debt_d = sum(
            ctx.vi.segment_net_debt[c]
            for c, info in ctx.vi.segments.items()
            if info.get("method") in ("pbv", "pe") and c in ctx.vi.segment_net_debt
        )
        eff_nd_d = ctx.vi.net_debt - fin_debt_d
        kpis.append(("순차입금 (연결)", ctx.vi.net_debt))
        kpis.append(("유효 순차입금 (제조)", eff_nd_d))
    else:
        kpis.append(("순차입금", ctx.vi.net_debt))
    kpis.append(("부채비율", f"{cons_by['de_ratio']:.1f}%"))
    if ctx.result.dcf:
        kpis.append(("DCF EV", ctx.result.dcf.ev_dcf))
    if ebitda > 0 and ctx.result.total_ev > 0:
        kpis.append(("EV/EBITDA (implied)", f"{ctx.result.total_ev / ebitda:.1f}x"))

    for label, val in kpis:
        write_cell(ws, r, 1, label)
        if isinstance(val, str):
            write_cell(ws, r, 2, val)
        else:
            write_cell(ws, r, 2, val, fmt=NUM_FMT)
        r += 1

    # ── Cross-validation ──
    cv_header_row = None
    if ctx.result.cross_validations:
        r += 1
        write_cell(ws, r, 1, "멀티플 교차검증 (Cross-Validation)", font=SECTION_FONT); r += 1
        cv_headers = ["방법론", "지표값", "배수", "EV", "Equity Value", f"주당 가치 ({ctx.currency_sym})"]
        cv_widths = [24, 18, 12, 20, 20, 20]
        for c, h in enumerate(cv_headers, 1):
            write_cell(ws, r, c, h)
            ws.column_dimensions[get_column_letter(c)].width = cv_widths[c - 1]
        style_header_row(ws, r, len(cv_headers))
        cv_header_row = r

        for cv in ctx.result.cross_validations:
            r += 1
            write_cell(ws, r, 1, cv.method)
            write_cell(ws, r, 2, cv.metric_value, fmt=NUM_FMT)
            write_cell(ws, r, 3, f"{cv.multiple:.1f}x" if cv.multiple > 0 else "-")
            write_cell(ws, r, 4, cv.enterprise_value, fmt=NUM_FMT)
            write_cell(ws, r, 5, cv.equity_value, fmt=NUM_FMT,
                       fill=GREEN_FILL if cv.equity_value > 0 else RED_FILL)
            write_cell(ws, r, 6, cv.per_share, fmt=NUM_FMT,
                       fill=GREEN_FILL if cv.per_share > 0 else RED_FILL)

    # ── SOTP vs DCF gap reconciliation note ──
    if ctx.result.cross_validations:
        sotp_val = next((cv.per_share for cv in ctx.result.cross_validations if cv.method == "SOTP"), None)
        dcf_val = next((cv.per_share for cv in ctx.result.cross_validations if cv.method == "DCF"), None)
        if sotp_val and dcf_val and min(sotp_val, dcf_val) > 0:
            gap_pct = abs(sotp_val - dcf_val) / min(sotp_val, dcf_val) * 100
            if gap_pct > 30:
                r += 1
                write_cell(ws, r, 1,
                    f"⚠ SOTP-DCF 괴리 {gap_pct:.0f}% — 비상장 복합기업 특성상 SOTP가 세그먼트별 시장 멀티플을 반영하는 반면, "
                    f"DCF는 보수적 FCFF 추정(높은 Capex·ΔNWC)을 적용하여 차이 발생. 투자 협상 가격 레인지 설정에 활용 가능.",
                    font=Font(italic=True, size=9, color="E74C3C"),
                )

    # ── Market price comparison ──
    if ctx.result.market_comparison and ctx.result.market_comparison.market_price > 0:
        mc = ctx.result.market_comparison
        r += 2
        write_cell(ws, r, 1, "시장가격 비교", font=SECTION_FONT); r += 1
        write_cell(ws, r, 1, "내재가치 (주당)")
        write_cell(ws, r, 2, mc.intrinsic_value, fmt=NUM_FMT); r += 1
        write_cell(ws, r, 1, "현재 시장가")
        write_cell(ws, r, 2, mc.market_price, fmt=NUM_FMT); r += 1
        write_cell(ws, r, 1, "괴리율")
        gap_fill = RED_FILL if abs(mc.gap_ratio) > 0.5 else GREEN_FILL
        write_cell(ws, r, 2, f"{mc.gap_ratio:.1%}", fill=gap_fill)
        if mc.flag:
            r += 1
            write_cell(ws, r, 1, mc.flag, font=Font(bold=True, color="E74C3C"))

    # ── Reverse-DCF Gap Diagnostic ──
    if ctx.result.gap_diagnostic:
        gd = ctx.result.gap_diagnostic
        _LABELS = {
            "wacc_overestimated": "WACC 과대추정",
            "growth_underestimated": "성장률 과소추정",
            "optionality_premium": "옵셔널리티 프리미엄",
            "market_pessimism": "시장 저평가 가능성",
        }
        r += 2
        write_cell(ws, r, 1, "역방향 DCF 진단", font=SECTION_FONT); r += 1
        write_cell(ws, r, 1, "진단 유형")
        cat_color = "E74C3C" if not gd.reconcilable else "2980B9"
        write_cell(ws, r, 2, _LABELS.get(gd.category, gd.category),
                   font=Font(bold=True, color=cat_color)); r += 1
        write_cell(ws, r, 1, "괴리율")
        write_cell(ws, r, 2, f"{abs(gd.gap_pct):.1f}% ({gd.direction})"); r += 1
        if gd.implied_wacc is not None:
            write_cell(ws, r, 1, "시장 내재 WACC")
            write_cell(ws, r, 2, f"{gd.implied_wacc:.2f}%"); r += 1
        if gd.implied_tgr is not None:
            write_cell(ws, r, 1, "시장 내재 TGR")
            write_cell(ws, r, 2, f"{gd.implied_tgr:.2f}%"); r += 1
        if gd.implied_growth_mult is not None:
            write_cell(ws, r, 1, "시장 내재 성장배수")
            write_cell(ws, r, 2, f"{gd.implied_growth_mult:.2f}x"); r += 1
        if not gd.reconcilable:
            write_cell(ws, r, 1, "⚠ EBITDA DCF로 설명 불가",
                       font=Font(bold=True, color="E74C3C")); r += 1
        if gd.suggestions:
            for s in gd.suggestions[:3]:
                write_cell(ws, r, 1, s, font=Font(italic=True, size=9)); r += 1
        if gd.explanation:
            write_cell(ws, r, 1, "상세 분석", font=Font(bold=True))
            write_cell(ws, r, 2, gd.explanation); r += 1

    # ── Monte Carlo ──
    if ctx.result.monte_carlo:
        mc = ctx.result.monte_carlo
        r += 2
        write_cell(ws, r, 1, f"Monte Carlo 시뮬레이션 ({mc.n_sims:,}회)", font=SECTION_FONT); r += 1
        
        # MC Input Assumptions
        write_cell(ws, r, 1, "MC 입력 분포 가정 (Input Assumptions)", font=Font(bold=True)); r += 1
        if mc.input_assumptions:
            for param, desc in mc.input_assumptions.items():
                write_cell(ws, r, 1, param)
                write_cell(ws, r, 2, desc)
                r += 1
        else:
            mc_assumptions = [
                ("멀티플 변동성 (Std Dev %)", f"{ctx.vi.mc_multiple_std_pct}%"),
                ("Revenue 변동성 (Std Dev %)", f"{ctx.vi.mc_revenue_std_pct}%"),
                ("DLOM 평균", f"{ctx.vi.mc_dlom_mean}%"),
                ("DLOM 변동성 (Std Dev %p)", f"{ctx.vi.mc_dlom_std}%p"),
            ]
            for label, val in mc_assumptions:
                write_cell(ws, r, 1, label)
                write_cell(ws, r, 2, val)
                r += 1
        r += 1

        mc_items = [
            ("Mean (평균)", mc.mean), ("Median (중위수)", mc.median),
            ("Std (표준편차)", mc.std),
            ("5th Percentile", mc.p5), ("25th Percentile", mc.p25),
            ("75th Percentile", mc.p75), ("95th Percentile", mc.p95),
            ("Min", mc.min_val), ("Max", mc.max_val),
        ]
        for label, val in mc_items:
            write_cell(ws, r, 1, label)
            fill = GREEN_FILL if label.startswith("Med") else None
            write_cell(ws, r, 2, val, fmt=NUM_FMT, fill=fill)
            write_cell(ws, r, 3, ctx.currency_sym)
            r += 1

        # Histogram
        if mc.histogram_bins:
            r += 1
            write_cell(ws, r, 1, "주당가치 분포 (히스토그램)", font=SECTION_FONT); r += 1
            write_cell(ws, r, 1, f"구간 ({ctx.currency_sym})")
            write_cell(ws, r, 2, "빈도")
            style_header_row(ws, r, 2)
            hist_start = r + 1
            for bin_val, cnt in zip(mc.histogram_bins, mc.histogram_counts):
                r += 1
                write_cell(ws, r, 1, bin_val, fmt=NUM_FMT)
                write_cell(ws, r, 2, cnt, fmt=NUM_FMT)
            hist_end = r

            hist_chart = BarChart()
            hist_chart.type = "col"
            hist_chart.style = 10
            hist_chart.title = "Monte Carlo — 주당가치 분포"
            hist_chart.y_axis.title = "빈도"
            hist_chart.x_axis.title = f"주당가치 ({ctx.currency_sym})"

            cats_h = Reference(ws, min_col=1, min_row=hist_start, max_row=hist_end)
            vals_h = Reference(ws, min_col=2, min_row=hist_start, max_row=hist_end)
            hist_chart.add_data(vals_h, titles_from_data=False)
            hist_chart.set_categories(cats_h)
            hist_chart.width = 22
            hist_chart.height = 13
            hist_chart.legend = None
            hist_chart.gapWidth = 10
            s_h = hist_chart.series[0]
            s_h.graphicalProperties.solidFill = "2E86C1"

            r += CHART_GAP
            ws.add_chart(hist_chart, f"A{r}")
            r += CHART_ROWS

    # ── Charts ──

    # Chart 1: Per-share value by scenario (only when scenarios exist)
    if sc_header_row and ctx.sc_codes:
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = f"시나리오별 주당 가치 ({ctx.currency_sym})"
        chart1.y_axis.title = ctx.currency_sym

        cats1 = Reference(ws, min_col=1, min_row=sc_header_row + 1, max_row=sc_header_row + len(ctx.sc_codes))
        vals1 = Reference(ws, min_col=2, min_row=sc_header_row, max_row=sc_header_row + len(ctx.sc_codes))
        chart1.add_data(vals1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.shape = 4
        chart1.width = 18
        chart1.height = 12

        sc_colors = ["1B2A4A", "27AE60", "E74C3C", "F39C12", "8E44AD"]
        s1 = chart1.series[0]
        s1.graphicalProperties.solidFill = sc_colors[0]
        for idx in range(1, len(ctx.sc_codes)):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = sc_colors[idx % len(sc_colors)]
            s1.data_points.append(pt)

        chart1.legend = None
        s1.dLbls = DataLabelList()
        s1.dLbls.showVal = True
        s1.dLbls.showSerName = False
        s1.dLbls.numFmt = '#,##0'

        r += CHART_GAP
        ws.add_chart(chart1, f"A{r}")
        r += CHART_ROWS

    # Chart 2: EV composition (SOTP only)
    if ev_data_start is not None and ev_data_end is not None and ev_data_end >= ev_data_start:
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = f"사업부별 Enterprise Value ({ctx.unit})"
        chart2.y_axis.title = ctx.unit

        cats2 = Reference(ws, min_col=1, min_row=ev_data_start, max_row=ev_data_end)
        vals2 = Reference(ws, min_col=2, min_row=ev_data_start, max_row=ev_data_end)
        chart2.add_data(vals2, titles_from_data=False)
        chart2.set_categories(cats2)
        chart2.shape = 4
        chart2.width = 18
        chart2.height = 12
        chart2.legend = None

        seg_colors = ["1B2A4A", "2E86C1", "27AE60", "F39C12", "8E44AD"]
        s2 = chart2.series[0]
        for idx, color in enumerate(seg_colors[:ev_data_end - ev_data_start + 1]):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = color
            s2.data_points.append(pt)

        s2.dLbls = DataLabelList()
        s2.dLbls.showVal = True
        s2.dLbls.showSerName = False
        s2.dLbls.numFmt = '#,##0'

        r += CHART_GAP
        ws.add_chart(chart2, f"A{r}")
        r += CHART_ROWS

    # ── Football Field ──
    r += CHART_GAP
    _write_football_field(ws, r, ctx)


def _get_primary_value(ctx: Ctx) -> tuple[int, str]:
    """Return the primary result value by valuation method."""
    if ctx.result.weighted_value > 0 and ctx.result.scenarios:
        return ctx.result.weighted_value, "확률가중 적정 주당 가치"

    if ctx.method == "ddm" and ctx.result.ddm:
        return ctx.result.ddm.equity_per_share, "DDM 적정 주당 가치"
    elif ctx.method == "rim" and ctx.result.rim:
        return ctx.result.rim.per_share, "RIM 적정 주당 가치"
    elif ctx.method == "nav" and ctx.result.nav:
        return ctx.result.nav.per_share, "NAV 적정 주당 가치"
    elif ctx.method == "multiples" and ctx.result.multiples_primary:
        return ctx.result.multiples_primary.per_share, "Multiples 적정 주당 가치"
    elif ctx.method == "dcf_primary" and ctx.result.dcf:
        return ctx.result.dcf.ev_dcf, f"DCF Enterprise Value ({ctx.unit})"
    else:
        return ctx.result.total_ev, f"Enterprise Value ({ctx.unit})"


def _write_football_field(ws, r: int, ctx: Ctx):
    """Football Field chart (based on cross-validation/scenarios)."""
    write_cell(ws, r, 1, "Football Field — 밸류에이션 범위", font=SECTION_FONT); r += 1
    FF_COL_START = 8
    ff_headers = ["방법론", "하단", "주당가치", "상단", "범위"]
    for c, h in enumerate(ff_headers):
        write_cell(ws, r, FF_COL_START + c, h)
    for col_idx in range(FF_COL_START, FF_COL_START + len(ff_headers)):
        cell = ws.cell(row=r, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = GRAY_FILL
    ff_header_row = r

    ff_colors_list = []
    ff_color_palette = ["1B2A4A", "2E86C1", "27AE60", "F39C12", "E74C3C", "8E44AD", "17A589"]

    FF_COL_LABEL = 8
    FF_COL_LO = 9
    FF_COL_VAL = 10
    FF_COL_HI = 11
    FF_COL_RANGE = 12

    ff_entries = []
    if ctx.result.cross_validations:
        for cv in ctx.result.cross_validations:
            ff_entries.append((cv.method, cv.per_share))
    elif ctx.result.scenarios:
        for sc_code in ctx.sc_codes:
            sr = ctx.result.scenarios[sc_code]
            sc = ctx.vi.scenarios[sc_code]
            ff_entries.append((f"{sc_code}: {sc.name}", sr.post_dlom))
    else:
        primary_val, primary_label = _get_primary_value(ctx)
        ff_entries.append((primary_label, primary_val))

    for i, (label, val) in enumerate(reversed(ff_entries)):
        lo = max(round(val * 0.8), 0)
        hi = round(val * 1.2) if val > 0 else 0

        r += 1
        write_cell(ws, r, FF_COL_LABEL, label)
        write_cell(ws, r, FF_COL_LO, lo, fmt=NUM_FMT)
        write_cell(ws, r, FF_COL_VAL, val, fmt=NUM_FMT, fill=BLUE_FILL)
        write_cell(ws, r, FF_COL_HI, hi, fmt=NUM_FMT)
        write_cell(ws, r, FF_COL_RANGE, max(hi - lo, 0), fmt=NUM_FMT)
        ff_colors_list.append(ff_color_palette[i % len(ff_color_palette)])

    ff_data_end = r
    ws.column_dimensions[get_column_letter(FF_COL_LABEL)].width = 26
    ws.column_dimensions[get_column_letter(FF_COL_LO)].width = 16
    ws.column_dimensions[get_column_letter(FF_COL_VAL)].width = 16
    ws.column_dimensions[get_column_letter(FF_COL_HI)].width = 16
    ws.column_dimensions[get_column_letter(FF_COL_RANGE)].width = 16

    if not ff_entries:
        return

    # Stacked bar chart
    chart3 = BarChart()
    chart3.type = "bar"
    chart3.style = 10
    chart3.title = "Football Field — 밸류에이션 범위"
    chart3.x_axis.numFmt = '#,##0'

    cats3 = Reference(ws, min_col=FF_COL_LABEL, min_row=ff_header_row + 1, max_row=ff_data_end)
    vals_lo = Reference(ws, min_col=FF_COL_LO, min_row=ff_header_row + 1, max_row=ff_data_end)
    vals_range = Reference(ws, min_col=FF_COL_RANGE, min_row=ff_header_row + 1, max_row=ff_data_end)

    chart3.add_data(vals_lo, titles_from_data=False)
    chart3.add_data(vals_range, titles_from_data=False)
    chart3.set_categories(cats3)
    chart3.grouping = "stacked"
    chart3.width = 20
    chart3.height = 10
    chart3.gapWidth = 80

    s3_lo = chart3.series[0]
    s3_lo.graphicalProperties.solidFill = "E8EAED"
    s3_lo.graphicalProperties.line.solidFill = "D5D8DC"

    s3_hi = chart3.series[1]
    for ci, color in enumerate(ff_colors_list):
        pt = DataPoint(idx=ci)
        pt.graphicalProperties.solidFill = color
        s3_hi.data_points.append(pt)

    s3_hi.dLbls = DataLabelList()
    s3_hi.dLbls.showVal = True
    s3_hi.dLbls.showSerName = False
    s3_hi.dLbls.showCatName = False
    s3_hi.dLbls.numFmt = '#,##0'
    chart3.legend = None

    r += 2
    ws.add_chart(chart3, f"A{r}")
