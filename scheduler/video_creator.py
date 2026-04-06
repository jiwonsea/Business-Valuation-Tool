"""Create weekly valuation report video (slides + TTS narration).

Design: Fey-inspired light theme — white background, card-based layout,
        dark text, green/red accents for financial data.
Charts: Extracted from Excel Dashboard via openpyxl → matplotlib → PNG.
Slides: Pillow (direct PNG, no LibreOffice dependency).
TTS:    Edge TTS (ko-KR-HyunsuNeural).
Video:  moviepy (image sequence + audio → MP4).

The TTS engine is abstracted into _synthesize() for future replacement
(e.g., GPT-SoVITS voice cloning).
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import re
import tempfile
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont

logger = logging.getLogger(__name__)

# ── Design Tokens (Fey-inspired light theme) ──

W, H = 1920, 1080
BG_WHITE = "#ffffff"
BG_CARD = "#f4f5f7"
BG_HEADER = "#0f172a"  # dark navy header bar
TEXT_PRIMARY = "#0f172a"  # near-black
TEXT_SECONDARY = "#64748b"  # slate gray
TEXT_ON_DARK = "#ffffff"
ACCENT_GREEN = "#22c55e"
ACCENT_RED = "#ef4444"
ACCENT_BLUE = "#3b82f6"
BORDER = "#e2e8f0"

TTS_VOICE = "ko-KR-HyunsuNeural"

_FONT_BOLD = "C:/Windows/Fonts/malgunbd.ttf"
_FONT_REGULAR = "C:/Windows/Fonts/malgun.ttf"


def _font_bold(size: int) -> ImageFont.FreeTypeFont:
    if os.path.exists(_FONT_BOLD):
        return ImageFont.truetype(_FONT_BOLD, size)
    return ImageFont.load_default()


def _font_regular(size: int) -> ImageFont.FreeTypeFont:
    if os.path.exists(_FONT_REGULAR):
        return ImageFont.truetype(_FONT_REGULAR, size)
    return ImageFont.load_default()


def _cap_str(cap: float | None) -> str:
    if not cap:
        return "N/A"
    if cap >= 1_000_000_000_000:
        return f"${cap / 1_000_000_000_000:.1f}T"
    if cap >= 1_000_000_000:
        return f"${cap / 1_000_000_000:.1f}B"
    return f"${cap / 1_000_000:,.0f}M"


def _draw_rounded_rect(
    draw: ImageDraw.ImageDraw, xy: tuple, radius: int, fill: str, outline: str | None = None,
) -> None:
    """Draw a rounded rectangle (card)."""
    x0, y0, x1, y1 = xy
    draw.rounded_rectangle(xy, radius=radius, fill=fill, outline=outline)


def _draw_header_bar(draw: ImageDraw.ImageDraw, text: str) -> None:
    """Draw a slim header bar at the top of the slide."""
    draw.rectangle([(0, 0), (W, 72)], fill=BG_HEADER)
    font = _font_bold(24)
    draw.text((60, 36), text, fill=TEXT_ON_DARK, font=font, anchor="lm")
    draw.text((W - 60, 36), "AI 자동 생성 · 투자 추천 아님", fill="#94a3b8", font=_font_regular(16), anchor="rm")


def _wrap_text(text: str, font: ImageFont.FreeTypeFont, max_width: int) -> list[str]:
    lines = []
    for paragraph in text.split("\n"):
        if not paragraph.strip():
            lines.append("")
            continue
        words = paragraph.split()
        if not words:
            lines.append("")
            continue
        current = words[0]
        for word in words[1:]:
            test = f"{current} {word}"
            bbox = font.getbbox(test)
            if bbox[2] - bbox[0] <= max_width:
                current = test
            else:
                lines.append(current)
                current = word
        lines.append(current)
    return lines


# ── Chart Extraction ──


def _extract_charts_from_excel(excel_path: str | Path) -> list[Image.Image]:
    """Extract Dashboard charts from Excel, recreate with matplotlib.

    Returns list of chart PNG images (as PIL Images).
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm
        import openpyxl
    except ImportError:
        logger.warning("matplotlib or openpyxl not available — skipping charts")
        return []

    excel_path = Path(excel_path)
    if not excel_path.exists():
        return []

    # Set Korean font for matplotlib
    font_path = _FONT_REGULAR
    if os.path.exists(font_path):
        fm.fontManager.addfont(font_path)
        prop = fm.FontProperties(fname=font_path)
        plt.rcParams["font.family"] = prop.get_name()
    plt.rcParams["axes.unicode_minus"] = False

    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    except Exception as e:
        logger.warning("Failed to open Excel %s: %s", excel_path, e)
        return []

    if "Dashboard" not in wb.sheetnames:
        return []

    ws = wb["Dashboard"]
    charts_data = []

    for chart in ws._charts:
        # Extract title
        title = ""
        try:
            for p in chart.title.tx.rich.paragraphs:
                for r in p.r:
                    title += r.t
        except Exception:
            title = "Chart"

        # Extract series data
        series_list = []
        for series in chart.series:
            cats = []
            vals = []

            # Categories (x-axis)
            if hasattr(series, "cat") and series.cat and series.cat.strRef:
                ref = series.cat.strRef.f
                cats = _read_range(ws, ref, wb)
            elif hasattr(series, "cat") and series.cat and series.cat.numRef:
                ref = series.cat.numRef.f
                cats = _read_range(ws, ref, wb)

            # Values
            if series.val and series.val.numRef:
                ref = series.val.numRef.f
                vals = _read_range(ws, ref, wb)

            if vals:
                series_title = ""
                if series.title and hasattr(series.title, "v") and series.title.v:
                    series_title = series.title.v
                series_list.append({"title": series_title, "cats": cats, "vals": vals})

        if series_list:
            charts_data.append({"title": title, "series": series_list})

    # Render charts with matplotlib
    chart_images = []
    for cdata in charts_data:
        try:
            fig, ax = plt.subplots(figsize=(8, 4), facecolor="white")
            ax.set_facecolor("#f8fafc")

            for s in cdata["series"]:
                vals = [v if isinstance(v, (int, float)) else 0 for v in s["vals"]]
                cats = s["cats"][:len(vals)] if s["cats"] else [str(i) for i in range(len(vals))]
                # Truncate labels
                cats = [str(c)[:12] if c else "" for c in cats]

                colors = [ACCENT_GREEN if v >= 0 else ACCENT_RED for v in vals]
                ax.bar(range(len(vals)), vals, color=colors, width=0.6, edgecolor="none")
                ax.set_xticks(range(len(cats)))
                ax.set_xticklabels(cats, rotation=30, ha="right", fontsize=9)

            ax.set_title(cdata["title"], fontsize=13, fontweight="bold", color=TEXT_PRIMARY, pad=12)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color(BORDER)
            ax.spines["bottom"].set_color(BORDER)
            ax.tick_params(colors=TEXT_SECONDARY, labelsize=9)
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:,.0f}"))
            plt.tight_layout()

            # Save to PIL Image
            fig.canvas.draw()
            chart_img = Image.frombytes(
                "RGB",
                fig.canvas.get_width_height(),
                fig.canvas.tostring_rgb(),
            )
            chart_images.append(chart_img)
            plt.close(fig)
        except Exception as e:
            logger.warning("Chart render failed (%s): %s", cdata["title"], e)
            plt.close("all")

    return chart_images


def _read_range(ws, ref: str, wb) -> list:
    """Read cell values from a sheet reference like \"'Dashboard'!B8:B12\"."""
    import re as _re

    match = _re.match(r"'?([^'!]+)'?!(.+)", ref)
    if not match:
        return []

    sheet_name, cell_range = match.groups()
    target_ws = wb[sheet_name] if sheet_name in wb.sheetnames else ws

    values = []
    try:
        for row in target_ws[cell_range]:
            if isinstance(row, tuple):
                for cell in row:
                    values.append(cell.value)
            else:
                values.append(row.value)
    except Exception:
        pass
    return values


# ── Slide Builders ──


def _slide_title(label: str) -> tuple[Image.Image, str]:
    """Title slide — clean white with centered title."""
    img = Image.new("RGB", (W, H), BG_WHITE)
    draw = ImageDraw.Draw(img)

    # Subtle top accent line
    draw.rectangle([(0, 0), (W, 6)], fill=ACCENT_BLUE)

    # Center content
    draw.text(
        (W // 2, 340),
        "주간 밸류에이션 리포트",
        fill=TEXT_PRIMARY,
        font=_font_bold(72),
        anchor="mm",
    )
    draw.text(
        (W // 2, 440),
        label,
        fill=ACCENT_BLUE,
        font=_font_regular(40),
        anchor="mm",
    )

    # Divider line
    draw.rectangle([(W // 2 - 100, 500), (W // 2 + 100, 502)], fill=BORDER)

    draw.text(
        (W // 2, 560),
        "AI Automated Valuation Pipeline",
        fill=TEXT_SECONDARY,
        font=_font_regular(24),
        anchor="mm",
    )

    draw.text(
        (W // 2, H - 50),
        "이 리포트는 AI가 자동 생성한 분석 자료이며 투자 추천이 아닙니다.",
        fill="#94a3b8",
        font=_font_regular(18),
        anchor="mm",
    )

    script = f"주간 밸류에이션 리포트. {label}."
    return img, script


def _slide_overview(summary: dict) -> tuple[Image.Image, str]:
    """Market overview slide with metric cards."""
    img = Image.new("RGB", (W, H), BG_WHITE)
    draw = ImageDraw.Draw(img)

    _draw_header_bar(draw, "시장 요약")

    discoveries = summary.get("discoveries", [])
    status = summary.get("status_summary", {})
    script_parts = ["이번 주 시장 요약입니다."]

    # Metric cards row
    card_y = 110
    card_h = 140
    card_w = (W - 180) // 3
    gap = 30

    metrics = [
        ("분석 완료", f"{status.get('success', 0)}개", ACCENT_GREEN),
        ("분석 대상", f"{status.get('total', 0)}개", ACCENT_BLUE),
        ("실패", f"{status.get('failed', 0)}개", ACCENT_RED if status.get("failed", 0) else TEXT_SECONDARY),
    ]

    for i, (label, value, color) in enumerate(metrics):
        x0 = 60 + i * (card_w + gap)
        _draw_rounded_rect(draw, (x0, card_y, x0 + card_w, card_y + card_h), 12, BG_CARD, BORDER)
        draw.text((x0 + 30, card_y + 30), label, fill=TEXT_SECONDARY, font=_font_regular(20))
        draw.text((x0 + 30, card_y + 70), value, fill=color, font=_font_bold(42))

    # Market discovery cards
    y = card_y + card_h + 40
    for d in discoveries:
        market = d.get("market", "")
        news_count = d.get("news_count", 0)
        cos = d.get("companies", [])
        co_names = ", ".join(c.get("name", "") for c in cos[:5])
        market_label = "[KR] 한국" if market == "KR" else "[US] 미국"

        _draw_rounded_rect(draw, (60, y, W - 60, y + 150), 12, BG_CARD, BORDER)
        draw.text((90, y + 20), market_label, fill=TEXT_PRIMARY, font=_font_bold(30))
        draw.text((90, y + 65), f"뉴스 {news_count}건 수집", fill=TEXT_SECONDARY, font=_font_regular(22))
        draw.text((90, y + 100), f"발굴: {co_names}", fill=TEXT_PRIMARY, font=_font_regular(22))

        y += 175

        script_parts.append(
            f"{market_label.split('] ')[1]} 시장에서는 뉴스 {news_count}건이 수집되었고, "
            f"주요 기업으로 {co_names}가 발굴되었습니다."
        )

    total = status.get("total", 0)
    success = status.get("success", 0)
    script_parts.append(f"총 {total}개 기업 중 {success}개의 분석이 완료되었습니다.")

    return img, " ".join(script_parts)


def _slide_company(entry: dict, chart_images: list[Image.Image] | None = None) -> tuple[Image.Image, str]:
    """Per-company slide with analysis text + optional charts."""
    img = Image.new("RGB", (W, H), BG_WHITE)
    draw = ImageDraw.Draw(img)

    name = entry.get("company", "Unknown")
    market = entry.get("market", "")
    cap = entry.get("market_cap_usd")
    summary_md = entry.get("summary_md", "")
    market_label = "한국" if market == "KR" else "미국"

    _draw_header_bar(draw, f"{name} — {market_label} | {_cap_str(cap)}")

    # Parse summary_md
    clean_lines = []
    for line in summary_md.strip().split("\n"):
        line = line.strip()
        if line.startswith("#"):
            continue
        line = re.sub(r"^\s*[-*]\s*", "• ", line)
        if line:
            clean_lines.append(line)

    has_chart = chart_images and len(chart_images) > 0

    if has_chart:
        # Left: text, Right: chart (2-column layout)
        text_width = W // 2 - 80
        text_x = 60

        # Chart area (right side)
        chart_x = W // 2 + 20
        chart_y = 100
        chart_w = W // 2 - 80
        chart_h = H - 180

        # Place up to 2 charts stacked
        for i, chart_img in enumerate(chart_images[:2]):
            target_h = chart_h // 2 - 10
            ratio = min(chart_w / chart_img.width, target_h / chart_img.height)
            new_w = int(chart_img.width * ratio)
            new_h = int(chart_img.height * ratio)
            resized = chart_img.resize((new_w, new_h), Image.LANCZOS)
            paste_y = chart_y + i * (target_h + 20)
            img.paste(resized, (chart_x + (chart_w - new_w) // 2, paste_y))
    else:
        text_width = W - 160
        text_x = 60

    # Render text
    body_font = _font_regular(24)
    y = 100
    for line in clean_lines[:14]:
        wrapped = _wrap_text(line, body_font, text_width)
        for wl in wrapped:
            if y > H - 80:
                break
            color = TEXT_PRIMARY
            if wl.startswith("•"):
                color = TEXT_PRIMARY
            draw.text((text_x, y), wl, fill=color, font=body_font)
            y += 36
        if y > H - 80:
            break

    # TTS script
    tts_lines = [f"{name}. {market_label} 시장. 시가총액 {_cap_str(cap)}."]
    for line in clean_lines[:8]:
        text = re.sub(r"[•#*\-]", "", line).strip()
        if text:
            tts_lines.append(text)

    return img, " ".join(tts_lines)


def _slide_closing() -> tuple[Image.Image, str]:
    """Closing slide."""
    img = Image.new("RGB", (W, H), BG_WHITE)
    draw = ImageDraw.Draw(img)

    draw.rectangle([(0, 0), (W, 6)], fill=ACCENT_BLUE)

    draw.text((W // 2, 380), "감사합니다", fill=TEXT_PRIMARY, font=_font_bold(64), anchor="mm")
    draw.text(
        (W // 2, 460),
        "매주 토요일 자동 업데이트됩니다",
        fill=TEXT_SECONDARY,
        font=_font_regular(28),
        anchor="mm",
    )

    draw.rectangle([(W // 2 - 80, 520), (W // 2 + 80, 522)], fill=BORDER)

    draw.text(
        (W // 2, H - 50),
        "이 영상은 AI가 자동 생성한 분석 자료이며 투자 추천이 아닙니다.",
        fill="#94a3b8",
        font=_font_regular(18),
        anchor="mm",
    )

    script = "이상으로 이번 주 밸류에이션 리포트를 마칩니다. 감사합니다."
    return img, script


# ── TTS Engine (swappable) ──


async def _synthesize(text: str, output_path: str) -> None:
    """Generate TTS audio using Edge TTS. Replace this function for voice cloning."""
    import edge_tts

    communicate = edge_tts.Communicate(text, TTS_VOICE)
    await communicate.save(output_path)


# ── Video Assembly ──


def create_weekly_video(summary: dict, output_dir: str | Path | None = None) -> Path | None:
    """Create weekly valuation report video.

    Args:
        summary: The full _weekly_summary.json content.
        output_dir: Directory to save the video. Defaults to valuation-results/{week_folder}.

    Returns:
        Path to the generated MP4 file, or None on failure.
    """
    from moviepy import AudioFileClip, ImageClip, concatenate_videoclips

    if output_dir is None:
        output_dir = Path(summary.get("week_dir", "test_output"))
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    label = summary.get("label", "Weekly Report")
    valuations = summary.get("valuations", [])
    success_entries = [v for v in valuations if v.get("status") == "success"]

    # Extract charts from Excel files (if available)
    company_charts: dict[str, list[Image.Image]] = {}
    for entry in success_entries:
        excel_path = entry.get("excel_path", "")
        if excel_path and Path(excel_path).exists():
            charts = _extract_charts_from_excel(excel_path)
            if charts:
                company_charts[entry["company"]] = charts

    # Build slides
    slides: list[tuple[Image.Image, str]] = []
    slides.append(_slide_title(label))
    slides.append(_slide_overview(summary))

    for entry in success_entries:
        charts = company_charts.get(entry.get("company", ""))
        slides.append(_slide_company(entry, chart_images=charts))

    slides.append(_slide_closing())

    # Generate TTS + video clips in temp dir
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        clips = []

        for i, (img, script) in enumerate(slides):
            img_path = tmp_path / f"slide_{i:02d}.png"
            audio_path = tmp_path / f"audio_{i:02d}.mp3"

            img.save(str(img_path))

            try:
                asyncio.run(_synthesize(script, str(audio_path)))
            except Exception as e:
                logger.error("TTS failed for slide %d: %s", i, e)
                continue

            try:
                audio_clip = AudioFileClip(str(audio_path))
                duration = max(audio_clip.duration + 0.5, 3.0)
                img_clip = (
                    ImageClip(str(img_path))
                    .with_duration(duration)
                    .with_audio(audio_clip)
                )
                clips.append(img_clip)
            except Exception as e:
                logger.error("Clip creation failed for slide %d: %s", i, e)
                continue

        if not clips:
            logger.error("No clips generated — video creation aborted.")
            return None

        video_path = output_dir / "weekly_video.mp4"
        try:
            final = concatenate_videoclips(clips, method="compose")
            final.write_videofile(
                str(video_path),
                fps=1,
                codec="libx264",
                audio_codec="aac",
                logger=None,
            )
            final.close()
            logger.info("Video created: %s (%.1f seconds)", video_path, final.duration)
            return video_path
        except Exception as e:
            logger.error("Video export failed: %s", e)
            return None


def main() -> None:
    """CLI entry point for standalone testing."""
    parser = argparse.ArgumentParser(description="Create weekly valuation video")
    parser.add_argument("--test", action="store_true", help="Test with latest/dummy summary JSON")
    parser.add_argument("--summary-json", type=str, help="Path to _weekly_summary.json")
    parser.add_argument("--output-dir", type=str, default="test_output", help="Output directory")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    if args.summary_json:
        summary_path = Path(args.summary_json)
    else:
        results_dir = Path(__file__).resolve().parent.parent / "valuation-results"
        summaries = sorted(results_dir.glob("*/_weekly_summary.json"), reverse=True)
        if summaries:
            summary_path = summaries[0]
        else:
            summary_path = Path(__file__).resolve().parent.parent / "test_output" / "_weekly_summary_dummy.json"

    if not summary_path.exists():
        logger.error("Summary file not found: %s", summary_path)
        return

    logger.info("Using summary: %s", summary_path)
    with open(summary_path, encoding="utf-8") as f:
        summary = json.load(f)

    video_path = create_weekly_video(summary, output_dir=args.output_dir)
    if video_path:
        logger.info("Video saved: %s", video_path)
    else:
        logger.error("Video creation failed.")


if __name__ == "__main__":
    main()
