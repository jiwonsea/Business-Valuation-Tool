"""Sheet 3: Valuation — method-specific (SOTP/DCF/DDM/RIM/NAV/Multiples)."""

from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from ._ctx import Ctx
from ..excel_styles import (
    BLUE_FILL, YELLOW_FILL, GREEN_FILL, RED_FILL, GRAY_FILL, DARK_FILL,
    SECTION_FONT, TITLE_FONT, NOTE_FONT, WHITE_FONT, RESULT_FONT,
    NUM_FMT, PCT_FMT, MULT_FMT,
    style_header_row, write_cell,
)


def valuation_sotp(ctx: Ctx):
    ws = ctx.wb.create_sheet("SOTP Valuation")
    ws.sheet_properties.tabColor = "27AE60"
    ws.column_dimensions['A'].width = 20

    write_cell(ws, 1, 1, f"SOTP 밸류에이션 ({ctx.by}년 기준, {ctx.unit})", font=TITLE_FONT)

    r = 3
    if ctx.result.sotp:
        # Check Mixed SOTP
        has_mixed = any(
            getattr(s, "method", "ev_ebitda") != "ev_ebitda"
            for s in ctx.result.sotp.values()
        )
        if has_mixed:
            write_cell(ws, r, 1, "부문별 SOTP (Mixed Method)", font=SECTION_FONT); r += 1
            sotp_headers = ["부문", "Method", "지표값", "멀티플", "Segment Value", "비중"]
            ncols = 6
        else:
            write_cell(ws, r, 1, "부문별 EV/EBITDA", font=SECTION_FONT); r += 1
            sotp_headers = ["부문", "EBITDA", "멀티플", "Segment EV", "EV 비중"]
            ncols = 5
        for c, h in enumerate(sotp_headers, 1):
            write_cell(ws, r, c, h)
            ws.column_dimensions[get_column_letter(c)].width = 16
        style_header_row(ws, r, ncols)

        for code in ctx.seg_codes:
            if code not in ctx.result.sotp:
                continue
            r += 1
            s = ctx.result.sotp[code]
            ev_pct = s.ev / ctx.result.total_ev if ctx.result.total_ev > 0 else 0
            if has_mixed:
                method = getattr(s, "method", "ev_ebitda")
                rev_type = getattr(s, "revenue_type", "ltm")
                rev_tag = f" ({rev_type.upper()})" if method == "ev_revenue" and rev_type != "ltm" else ""
                method_label = {"ev_ebitda": "EV/EBITDA", "pbv": "P/BV", "pe": "P/E", "ev_revenue": "EV/Revenue"}.get(method, method) + rev_tag
                seg_info = ctx.vi.segments.get(code, {})
                if method == "pbv":
                    metric_val = seg_info.get("book_equity", 0)
                elif method == "pe":
                    metric_val = seg_info.get("net_income_segment", 0)
                elif method == "ev_revenue":
                    metric_val = getattr(s, "revenue", 0) or 0
                else:
                    metric_val = s.ebitda
                write_cell(ws, r, 1, ctx.seg_names[code])
                write_cell(ws, r, 2, method_label)
                write_cell(ws, r, 3, metric_val, fmt=NUM_FMT)
                write_cell(ws, r, 4, s.multiple, fmt=MULT_FMT, fill=BLUE_FILL)
                write_cell(ws, r, 5, s.ev, fmt=NUM_FMT, fill=GREEN_FILL if s.ev > 0 else None)
                write_cell(ws, r, 6, ev_pct, fmt=PCT_FMT)
            else:
                write_cell(ws, r, 1, ctx.seg_names[code])
                write_cell(ws, r, 2, s.ebitda, fmt=NUM_FMT)
                write_cell(ws, r, 3, s.multiple, fmt=MULT_FMT, fill=BLUE_FILL)
                write_cell(ws, r, 4, s.ev, fmt=NUM_FMT, fill=GREEN_FILL if s.ev > 0 else None)
                write_cell(ws, r, 5, ev_pct, fmt=PCT_FMT)
            # Negative EBITDA warning
            if s.ebitda <= 0 and getattr(s, "method", "ev_ebitda") == "ev_ebitda":
                r += 1
                write_cell(ws, r, 1,
                    f"  ⚠ {ctx.seg_names[code]}: EBITDA ≤ 0 → EV/EBITDA 무의미. EV=0 처리 (보수적 가정). EV/Revenue 또는 청산가치 대안 검토 필요.",
                    font=Font(italic=True, size=9, color="E74C3C"),
                )
        r += 1
        write_cell(ws, r, 1, "합계", bold=True)
        if has_mixed:
            write_cell(ws, r, 5, ctx.result.total_ev, fmt=NUM_FMT, bold=True, fill=GREEN_FILL)
            write_cell(ws, r, 6, 1.0, fmt=PCT_FMT, bold=True)
        else:
            write_cell(ws, r, 2, sum(ctx.result.sotp[c].ebitda for c in ctx.seg_codes if c in ctx.result.sotp), fmt=NUM_FMT, bold=True)
            write_cell(ws, r, 4, ctx.result.total_ev, fmt=NUM_FMT, bold=True, fill=GREEN_FILL)
            write_cell(ws, r, 5, 1.0, fmt=PCT_FMT, bold=True)

        # Mixed SOTP: display Equity Bridge
        if has_mixed:
            r += 2
            write_cell(ws, r, 1, "Equity Bridge (Mixed SOTP)", font=SECTION_FONT); r += 1
            ev_segs_val = sum(s.ev for s in ctx.result.sotp.values() if not getattr(s, "is_equity_based", False))
            eq_segs_val = sum(s.ev for s in ctx.result.sotp.values() if getattr(s, "is_equity_based", False))
            fin_debt = sum(
                ctx.vi.segment_net_debt.get(c, 0)
                for c, info in ctx.vi.segments.items()
                if info.get("method") in ("pbv", "pe")
            )
            eff_nd = ctx.vi.net_debt - fin_debt
            bridge_items = [
                ("제조 부문 EV (EV/EBITDA)", ev_segs_val),
                ("(-) 유효 순차입금 (제조)", eff_nd),
                ("제조 부문 Equity", ev_segs_val - eff_nd),
                ("(+) 금융 부문 Equity (P/BV)", eq_segs_val),
                ("Total Equity", ev_segs_val - eff_nd + eq_segs_val),
            ]
            for label, val in bridge_items:
                is_total = label == "Total Equity"
                write_cell(ws, r, 1, label, bold=is_total)
                write_cell(ws, r, 2, val, fmt=NUM_FMT, bold=is_total,
                           fill=GREEN_FILL if is_total else YELLOW_FILL)
                write_cell(ws, r, 3, ctx.unit)
                r += 1


def valuation_dcf(ctx: Ctx):
    ws = ctx.wb.create_sheet("DCF Valuation")
    ws.sheet_properties.tabColor = "2E86C1"
    ws.column_dimensions['A'].width = 24

    write_cell(ws, 1, 1, f"DCF 밸류에이션 — FCFF ({ctx.unit})", font=TITLE_FONT)

    dcf = ctx.result.dcf
    if not dcf:
        write_cell(ws, 3, 1, "DCF 결과 없음", font=SECTION_FONT)
        return

    # FCF Projection table
    r = 3
    write_cell(ws, r, 1, "Free Cash Flow 추정", font=SECTION_FONT); r += 1
    proj_headers = ["연도", "EBITDA", "D&A", "영업이익", "NOPAT", "Capex", "ΔNWC", "FCFF", "성장률", "PV(FCFF)"]
    for c, h in enumerate(proj_headers, 1):
        write_cell(ws, r, c, h)
        ws.column_dimensions[get_column_letter(c)].width = 14
    style_header_row(ws, r, len(proj_headers))

    for p in dcf.projections:
        r += 1
        vals = [
            (p.year, None), (p.ebitda, NUM_FMT), (p.da, NUM_FMT), (p.op, NUM_FMT),
            (p.nopat, NUM_FMT), (p.capex, NUM_FMT), (p.delta_nwc, NUM_FMT),
            (p.fcff, NUM_FMT), (p.growth, PCT_FMT), (p.pv_fcff, NUM_FMT),
        ]
        for c, (v, fmt) in enumerate(vals, 1):
            fill = GREEN_FILL if c == 8 and v > 0 else (RED_FILL if c == 8 and v < 0 else None)
            write_cell(ws, r, c, v, fmt=fmt, fill=fill)

    # DCF summary
    r += 2
    write_cell(ws, r, 1, "DCF 밸류에이션 요약", font=SECTION_FONT); r += 1
    summary = [
        ("PV(FCFF) 합계", dcf.pv_fcff_sum),
        ("Terminal Value", dcf.terminal_value),
        ("PV(Terminal Value)", dcf.pv_terminal),
        ("Enterprise Value (DCF)", dcf.ev_dcf),
    ]
    for label, val in summary:
        write_cell(ws, r, 1, label)
        is_ev = "Enterprise" in label
        write_cell(ws, r, 2, val, fmt=NUM_FMT,
                   fill=GREEN_FILL if is_ev else YELLOW_FILL,
                   bold=is_ev)
        write_cell(ws, r, 3, ctx.unit)
        r += 1

    r += 1
    write_cell(ws, r, 1, f"WACC: {dcf.wacc:.2f}%  |  Terminal Growth: {dcf.terminal_growth:.1f}%",
               font=NOTE_FONT)


def valuation_ddm(ctx: Ctx):
    ws = ctx.wb.create_sheet("DDM Valuation")
    ws.sheet_properties.tabColor = "27AE60"
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 36

    write_cell(ws, 1, 1, "DDM 밸류에이션 — 배당할인모델 (Gordon Growth)", font=TITLE_FONT)

    ddm = ctx.result.ddm
    if not ddm:
        write_cell(ws, 3, 1, "DDM 결과 없음", font=SECTION_FONT)
        return

    r = 3
    write_cell(ws, r, 1, "DDM 핵심 파라미터", font=SECTION_FONT); r += 1
    ddm_items = [
        ("주당 배당금 (DPS)", f"{ddm.dps:,.0f}", "최근 실적 기반"),
    ]
    if ddm.buyback_per_share > 0:
        ddm_items.append(("주당 자사주매입", f"{ddm.buyback_per_share:,.0f}", "Total Payout"))
        ddm_items.append(("Total Payout/주", f"{ddm.total_payout:,.0f}", "DPS + Buyback"))
    ddm_items += [
        ("배당 성장률 (g)", f"{ddm.growth:.2f}%", "지속가능 성장률"),
        ("자기자본비용 (Ke)", f"{ddm.ke:.2f}%", "CAPM: Rf + βL × ERP"),
        ("", "", ""),
        ("주당 내재가치", f"{ddm.equity_per_share:,}", "DPS×(1+g) / (Ke-g)"),
    ]
    for label, val, note in ddm_items:
        if not label:
            r += 1; continue
        write_cell(ws, r, 1, label)
        is_result = "내재가치" in label
        fill = GREEN_FILL if is_result else BLUE_FILL
        font = RESULT_FONT if is_result else None
        write_cell(ws, r, 2, val, fill=fill, font=font)
        write_cell(ws, r, 3, note)
        r += 1

    # DDM sensitivity (Ke x Growth)
    r += 1
    _write_ddm_sensitivity(ws, r, ddm, ctx.currency_sym)


def _write_ddm_sensitivity(ws, r: int, ddm, currency_sym: str):
    """DDM Ke x Growth sensitivity table."""
    write_cell(ws, r, 1, f"DDM 민감도 — Ke × 배당성장률 → 주당가치 ({currency_sym})", font=SECTION_FONT)
    r += 1

    ke_base, g_base = ddm.ke, ddm.growth
    ke_range = [ke_base + d for d in [-2.0, -1.0, -0.5, 0.0, 0.5, 1.0, 2.0]]
    g_range = [g_base + d for d in [-1.5, -1.0, -0.5, 0.0, 0.5, 1.0, 1.5]]

    write_cell(ws, r, 1, "Ke \\ Growth", fill=GRAY_FILL, font=SECTION_FONT)
    for j, g_val in enumerate(g_range, 2):
        write_cell(ws, r, j, f"{g_val:.1f}%", fill=GRAY_FILL, font=SECTION_FONT)
        ws.column_dimensions[get_column_letter(j)].width = 12

    from engine.ddm import calc_ddm as _calc_ddm
    sens_start = r + 1
    for ke_val in ke_range:
        r += 1
        write_cell(ws, r, 1, f"{ke_val:.1f}%", fill=GRAY_FILL, font=SECTION_FONT)
        for j, g_val in enumerate(g_range, 2):
            try:
                v = _calc_ddm(ddm.dps, g_val, ke_val,
                              buyback_per_share=getattr(ddm, 'buyback_per_share', 0.0)).equity_per_share
            except ValueError:
                v = 0
            is_base = abs(ke_val - ke_base) < 0.01 and abs(g_val - g_base) < 0.01
            fill = GREEN_FILL if is_base else (RED_FILL if v <= 0 else None)
            write_cell(ws, r, j, v, fmt=NUM_FMT, fill=fill)
    sens_end = r

    if g_range:
        end_col = get_column_letter(1 + len(g_range))
        ws.conditional_formatting.add(
            f"B{sens_start}:{end_col}{sens_end}", ColorScaleRule(
            start_type='min', start_color='FADBD8',
            mid_type='percentile', mid_value=50, mid_color='F5F6FA',
            end_type='max', end_color='D5F5E3',
        ))


def valuation_rim(ctx: Ctx):
    ws = ctx.wb.create_sheet("RIM Valuation")
    ws.sheet_properties.tabColor = "8E44AD"
    for col in 'ABCDEF':
        ws.column_dimensions[col].width = 20
    ws.column_dimensions['A'].width = 28

    write_cell(ws, 1, 1, "RIM 밸류에이션 — 잔여이익모델 (Residual Income)", font=TITLE_FONT)

    rim = ctx.result.rim
    if not rim:
        write_cell(ws, 3, 1, "RIM 결과 없음", font=SECTION_FONT)
        return

    r = 3
    write_cell(ws, r, 1, "RIM 핵심 파라미터", font=SECTION_FONT); r += 1
    rim_items = [
        ("장부가치 (BV₀)", f"{rim.bv_current:,}", "현재 자기자본"),
        ("자기자본비용 (Ke)", f"{rim.ke:.2f}%", "CAPM: Rf + βL × ERP"),
        ("영구성장률 (g)", f"{rim.terminal_growth:.2f}%", "RI Terminal Growth"),
        ("PV(RI)", f"{rim.pv_ri_sum:,}", "예측기간 잔여이익 현재가치"),
        ("PV(TV)", f"{rim.pv_terminal:,}", "잔여이익 Terminal Value 현재가치"),
        ("자기자본가치", f"{rim.equity_value:,}", "BV + PV(RI) + PV(TV)"),
        ("주당 내재가치", f"{rim.per_share:,}", "자기자본가치 / 주식수"),
    ]
    for label, val, note in rim_items:
        write_cell(ws, r, 1, label, fill=GRAY_FILL, font=WHITE_FONT)
        is_result = "내재가치" in label
        write_cell(ws, r, 2, val, fill=GREEN_FILL if is_result else None,
                   font=RESULT_FONT if is_result else None)
        write_cell(ws, r, 3, note, font=NOTE_FONT)
        r += 1

    # Year-by-year projections
    r += 1
    write_cell(ws, r, 1, "연도별 잔여이익 예측", font=SECTION_FONT); r += 1
    headers = ["Year", "기초 BV", "당기순이익", "ROE (%)", "잔여이익 (RI)", "PV(RI)"]
    for j, h in enumerate(headers, 1):
        write_cell(ws, r, j, h, fill=DARK_FILL, font=WHITE_FONT)
    r += 1
    for p in rim.projections:
        write_cell(ws, r, 1, f"Y{p.year}")
        write_cell(ws, r, 2, p.bv, fmt=NUM_FMT)
        write_cell(ws, r, 3, p.net_income, fmt=NUM_FMT)
        write_cell(ws, r, 4, p.roe / 100, fmt=PCT_FMT)
        write_cell(ws, r, 5, p.ri, fmt=NUM_FMT)
        write_cell(ws, r, 6, p.pv_ri, fmt=NUM_FMT)
        r += 1


def valuation_nav(ctx: Ctx):
    ws = ctx.wb.create_sheet("NAV Valuation")
    ws.sheet_properties.tabColor = "E67E22"
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 36

    write_cell(ws, 1, 1, "NAV 밸류에이션 — 순자산가치평가법", font=TITLE_FONT)

    nav = ctx.result.nav
    if not nav:
        write_cell(ws, 3, 1, "NAV 결과 없음", font=SECTION_FONT)
        return

    r = 3
    write_cell(ws, r, 1, f"순자산가치 구성 ({ctx.unit})", font=SECTION_FONT); r += 1
    nav_items = [
        ("총자산 (장부가)", nav.total_assets, ""),
        ("(+) 투자자산 재평가", nav.revaluation, "공정가치 − 장부가"),
        ("조정 후 총자산", nav.adjusted_assets, ""),
        ("(-) 총부채", nav.total_liabilities, ""),
        ("", 0, ""),
        ("순자산가치 (NAV)", nav.nav, "조정자산 − 부채"),
        ("주당 NAV", nav.per_share, f"{ctx.currency_sym}"),
    ]
    for label, val, note in nav_items:
        if not label:
            r += 1; continue
        write_cell(ws, r, 1, label)
        is_result = "주당" in label or "순자산가치 (NAV)" in label
        write_cell(ws, r, 2, val, fmt=NUM_FMT,
                   fill=GREEN_FILL if is_result else YELLOW_FILL,
                   font=RESULT_FONT if "주당" in label else None)
        if note:
            write_cell(ws, r, 3, note, font=NOTE_FONT)
        r += 1


def valuation_multiples(ctx: Ctx):
    ws = ctx.wb.create_sheet("Multiples Valuation")
    ws.sheet_properties.tabColor = "17A589"
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    write_cell(ws, 1, 1, "상대가치평가 — Multiples Primary", font=TITLE_FONT)

    mp = ctx.result.multiples_primary
    if not mp:
        write_cell(ws, 3, 1, "Multiples 결과 없음", font=SECTION_FONT)
        return

    r = 3
    write_cell(ws, r, 1, "적용 방법론", font=SECTION_FONT); r += 1
    items = [
        ("방법론", mp.primary_multiple_method, ""),
        ("지표값", f"{mp.metric_value:,.0f}", ctx.unit),
        ("적용 멀티플", f"{mp.multiple:.1f}x", "Peer Median 기반"),
        ("Enterprise Value", f"{mp.enterprise_value:,}", ctx.unit),
        ("Equity Value", f"{mp.equity_value:,}", ctx.unit),
        ("주당 가치", f"{mp.per_share:,}", ctx.currency_sym),
    ]
    for label, val, note in items:
        write_cell(ws, r, 1, label)
        is_result = "주당" in label
        write_cell(ws, r, 2, val, fill=GREEN_FILL if is_result else BLUE_FILL,
                   font=RESULT_FONT if is_result else None)
        if note:
            write_cell(ws, r, 3, note, font=NOTE_FONT)
        r += 1


VALUATION_MAP = {
    "sotp": valuation_sotp,
    "dcf_primary": valuation_dcf,
    "ddm": valuation_ddm,
    "rim": valuation_rim,
    "nav": valuation_nav,
    "multiples": valuation_multiples,
    # "rnpv" is handled separately in excel_builder.py (two sheets)
}
