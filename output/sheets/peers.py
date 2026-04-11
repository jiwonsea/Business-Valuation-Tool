"""Sheet 4: Peer Comparison — comparable company analysis."""

from openpyxl.utils import get_column_letter

from ._ctx import Ctx
from ..excel_styles import (
    YELLOW_FILL,
    GREEN_FILL,
    SECTION_FONT,
    TITLE_FONT,
    MULT_FMT,
    style_header_row,
    write_cell,
)


def sheet_peers(ctx: Ctx):
    if not ctx.vi.peers and not ctx.result.peer_stats:
        return

    ws = ctx.wb.create_sheet("Peer Comparison")
    ws.sheet_properties.tabColor = "17A589"
    write_cell(
        ws, 1, 1, "유사기업 비교분석 (Comparable Company Analysis)", font=TITLE_FONT
    )

    r = 3
    has_extra = any(p.ticker for p in ctx.vi.peers)
    if has_extra:
        peer_headers = [
            "기업명",
            "Ticker",
            "매핑 부문",
            "EV/EBITDA",
            "P/E (TTM)",
            "P/BV",
            "Beta",
            "출처",
            "비고",
        ]
        col_widths = [20, 10, 16, 12, 12, 10, 8, 8, 40]
    else:
        peer_headers = ["기업명", "매핑 부문", "EV/EBITDA", "비고"]
        col_widths = [20, 18, 12, 50]
    for c, h in enumerate(peer_headers, 1):
        write_cell(ws, r, c, h)
        ws.column_dimensions[get_column_letter(c)].width = col_widths[c - 1]
    style_header_row(ws, r, len(peer_headers))

    for p in ctx.vi.peers:
        r += 1
        c = 1
        write_cell(ws, r, c, p.name)
        c += 1
        if has_extra:
            write_cell(ws, r, c, p.ticker or "-")
            c += 1
        write_cell(ws, r, c, ctx.seg_names.get(p.segment_code, p.segment_code))
        c += 1
        write_cell(ws, r, c, p.ev_ebitda, fmt=MULT_FMT, fill=YELLOW_FILL)
        c += 1
        if has_extra:
            write_cell(
                ws, r, c, p.trailing_pe or "-", fmt=MULT_FMT if p.trailing_pe else None
            )
            c += 1
            write_cell(ws, r, c, p.pbv or "-", fmt=MULT_FMT if p.pbv else None)
            c += 1
            write_cell(ws, r, c, f"{p.beta:.2f}" if p.beta else "-")
            c += 1
            write_cell(ws, r, c, p.source)
            c += 1
        write_cell(ws, r, c, p.notes)

    # Per-segment multiple statistics
    if ctx.result.peer_stats:
        r += 2
        write_cell(ws, r, 1, "부문별 EV/EBITDA 멀티플 통계", font=SECTION_FONT)
        r += 1
        stat_headers = [
            "부문",
            "Peer 수",
            "Min",
            "Q1",
            "Median",
            "Mean",
            "Q3",
            "Max",
            "적용 멀티플",
        ]
        for c, h in enumerate(stat_headers, 1):
            write_cell(ws, r, c, h)
            ws.column_dimensions[get_column_letter(c)].width = max(
                ws.column_dimensions[get_column_letter(c)].width or 0,
                [18, 8, 8, 8, 8, 8, 8, 8, 12][c - 1],
            )
        style_header_row(ws, r, len(stat_headers))

        for ps in ctx.result.peer_stats:
            r += 1
            write_cell(ws, r, 1, ps.segment_name)
            write_cell(ws, r, 2, ps.count)
            write_cell(ws, r, 3, ps.ev_ebitda_min, fmt=MULT_FMT)
            write_cell(ws, r, 4, ps.ev_ebitda_q1, fmt=MULT_FMT)
            write_cell(ws, r, 5, ps.ev_ebitda_median, fmt=MULT_FMT, fill=GREEN_FILL)
            write_cell(ws, r, 6, ps.ev_ebitda_mean, fmt=MULT_FMT)
            write_cell(ws, r, 7, ps.ev_ebitda_q3, fmt=MULT_FMT)
            write_cell(ws, r, 8, ps.ev_ebitda_max, fmt=MULT_FMT)
            applied = ps.applied_multiple
            fill = (
                GREEN_FILL if abs(applied - ps.ev_ebitda_median) <= 2.0 else YELLOW_FILL
            )
            write_cell(ws, r, 9, applied, fmt=MULT_FMT, fill=fill, bold=True)
