"""Sheet 2: Financial Summary — consolidated financials + per-segment D&A."""

from openpyxl.utils import get_column_letter

from ._ctx import Ctx
from ..excel_styles import (
    YELLOW_FILL, GREEN_FILL, RED_FILL,
    SECTION_FONT, TITLE_FONT,
    NUM_FMT, PCT_FMT,
    style_header_row, write_cell,
)


def sheet_financials(ctx: Ctx):
    ws = ctx.wb.create_sheet("Financial Summary")
    ws.sheet_properties.tabColor = "2E86C1"
    ws.column_dimensions['A'].width = 24

    write_cell(ws, 1, 1, f"연결 재무제표 요약 ({ctx.unit})", font=TITLE_FONT)

    headers = ["항목"] + [str(y) for y in ctx.years]
    r = 3
    for c, h in enumerate(headers, 1):
        write_cell(ws, r, c, h)
        ws.column_dimensions[get_column_letter(c)].width = 18 if c > 1 else 24
    style_header_row(ws, r, len(headers))

    cons = ctx.cons
    years = ctx.years
    rows_data = [
        ("매출액", [cons[y]["revenue"] for y in years]),
        ("영업이익", [cons[y]["op"] for y in years]),
        ("당기순이익", [cons[y]["net_income"] for y in years]),
        ("총자산", [cons[y]["assets"] for y in years]),
        ("총부채", [cons[y]["liabilities"] for y in years]),
        ("총자본", [cons[y]["equity"] for y in years]),
        ("부채비율 (%)", [cons[y]["de_ratio"] for y in years]),
        ("감가상각비", [cons[y]["dep"] for y in years]),
        ("무형자산상각비", [cons[y]["amort"] for y in years]),
        ("D&A 합계", [cons[y]["dep"] + cons[y]["amort"] for y in years]),
        ("EBITDA", [cons[y]["op"] + cons[y]["dep"] + cons[y]["amort"] for y in years]),
    ]
    for label, vals in rows_data:
        r += 1
        write_cell(ws, r, 1, label, bold=True)
        for i, v in enumerate(vals, 2):
            f = '#,##0' if isinstance(v, int) else '0.0'
            write_cell(ws, r, i, v, fmt=f, fill=YELLOW_FILL)

    # Per-segment D&A allocation (SOTP only)
    if ctx.method == "sotp" and ctx.result.da_allocations:
        r += 2
        # Check Mixed SOTP
        _has_mixed_fs = any(
            info.get("method") in ("pbv", "pe") for info in ctx.vi.segments.values()
        )
        fs_title = "부문별 재무 — D&A 배분 (금융 부문 제외)" if _has_mixed_fs else "부문별 재무 — 유무형자산 비중 D&A 배분"
        write_cell(ws, r, 1, fs_title, font=TITLE_FONT); r += 1

        for yr in reversed(years):
            if yr not in ctx.result.da_allocations:
                continue
            r += 1
            write_cell(ws, r, 1, f"── {yr}년 ──", font=SECTION_FONT); r += 1
            if _has_mixed_fs:
                seg_headers = ["부문", "Method", "매출", "영업이익", "유무형자산", "자산비중", "D&A 배분", "EBITDA"]
                ncols_fs = 8
            else:
                seg_headers = ["부문", "매출", "영업이익", "유무형자산", "자산비중", "D&A 배분", "EBITDA"]
                ncols_fs = 7
            _seg_width = {"부문": 24, "Method": 14, "자산비중": 12}
            for c, h in enumerate(seg_headers, 1):
                write_cell(ws, r, c, h)
                ws.column_dimensions[get_column_letter(c)].width = _seg_width.get(h, 18)
            style_header_row(ws, r, ncols_fs)

            alloc = ctx.result.da_allocations[yr]
            for code in ctx.seg_codes:
                r += 1
                s = ctx.vi.segment_data[yr][code]
                a = alloc[code]
                if _has_mixed_fs:
                    seg_method = ctx.vi.segments.get(code, {}).get("method", "ev_ebitda")
                    m_lbl = {"ev_ebitda": "EV/EBITDA", "pbv": "P/BV", "pe": "P/E", "ev_revenue": "EV/Revenue"}.get(seg_method, seg_method)
                    write_cell(ws, r, 1, ctx.seg_names[code])
                    write_cell(ws, r, 2, m_lbl)
                    write_cell(ws, r, 3, s["revenue"], fmt=NUM_FMT, fill=YELLOW_FILL)
                    write_cell(ws, r, 4, s["op"], fmt=NUM_FMT, fill=YELLOW_FILL)
                    write_cell(ws, r, 5, s["assets"], fmt=NUM_FMT, fill=YELLOW_FILL)
                    write_cell(ws, r, 6, a.asset_share / 100, fmt=PCT_FMT)
                    write_cell(ws, r, 7, a.da_allocated, fmt=NUM_FMT)
                    write_cell(ws, r, 8, a.ebitda, fmt=NUM_FMT,
                               fill=GREEN_FILL if a.ebitda > 0 else RED_FILL)
                else:
                    write_cell(ws, r, 1, ctx.seg_names[code])
                    write_cell(ws, r, 2, s["revenue"], fmt=NUM_FMT, fill=YELLOW_FILL)
                    write_cell(ws, r, 3, s["op"], fmt=NUM_FMT, fill=YELLOW_FILL)
                    write_cell(ws, r, 4, s["assets"], fmt=NUM_FMT, fill=YELLOW_FILL)
                    write_cell(ws, r, 5, a.asset_share / 100, fmt=PCT_FMT)
                    write_cell(ws, r, 6, a.da_allocated, fmt=NUM_FMT)
                    write_cell(ws, r, 7, a.ebitda, fmt=NUM_FMT,
                               fill=GREEN_FILL if a.ebitda > 0 else RED_FILL)

            r += 1
            write_cell(ws, r, 1, "합계", bold=True)
            if _has_mixed_fs:
                write_cell(ws, r, 3, sum(ctx.vi.segment_data[yr][c]["revenue"] for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
                write_cell(ws, r, 4, sum(ctx.vi.segment_data[yr][c]["op"] for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
                write_cell(ws, r, 5, sum(ctx.vi.segment_data[yr][c]["assets"] for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
                write_cell(ws, r, 6, 1.0, fmt=PCT_FMT, bold=True)
                write_cell(ws, r, 7, sum(alloc[c].da_allocated for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
                write_cell(ws, r, 8, sum(alloc[c].ebitda for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
            else:
                write_cell(ws, r, 2, sum(ctx.vi.segment_data[yr][c]["revenue"] for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
                write_cell(ws, r, 3, sum(ctx.vi.segment_data[yr][c]["op"] for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
                write_cell(ws, r, 4, sum(ctx.vi.segment_data[yr][c]["assets"] for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
                write_cell(ws, r, 5, 1.0, fmt=PCT_FMT, bold=True)
                write_cell(ws, r, 6, sum(alloc[c].da_allocated for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
                write_cell(ws, r, 7, sum(alloc[c].ebitda for c in ctx.seg_codes), fmt=NUM_FMT, bold=True)
