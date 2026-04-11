"""Sheet 6: Sensitivity — method-specific sensitivity tables + heatmaps."""

from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font

from ._ctx import Ctx
from ..excel_styles import (
    GRAY_FILL, GREEN_FILL, RED_FILL,
    HEADER_FONT, SECTION_FONT, TITLE_FONT,
    NUM_FMT,
    write_cell,
)


def sheet_sensitivity(ctx: Ctx):
    ws = ctx.wb.create_sheet("Sensitivity")
    ws.sheet_properties.tabColor = "E74C3C"
    write_cell(ws, 1, 1, "민감도 분석", font=TITLE_FONT)

    r = 3
    method = ctx.method

    # ── SOTP: multiple x multiple ──
    if method == "sotp" and ctx.result.sensitivity_multiples:
        row_name = ctx.seg_names.get(ctx.seg_codes[0], "Row") if len(ctx.seg_codes) > 0 else "Row"
        col_name = ctx.seg_names.get(ctx.seg_codes[1], "Col") if len(ctx.seg_codes) > 1 else row_name
        r = _write_sensitivity_table(
            ws, r,
            f"① 멀티플 민감도 → 주당가치 ({ctx.currency_sym})",
            ctx.result.sensitivity_multiples,
            f"{row_name} \\ {col_name}", lambda v: f"{v:.0f}x", lambda v: f"{v:.0f}x",
        )
        r += 2

    # ── IRR x DLOM (private companies only) ──
    if ctx.result.sensitivity_irr_dlom:
        r = _write_sensitivity_table(
            ws, r,
            f"{'② ' if method == 'sotp' else '① '}FI IRR × DLOM → 주당가치 ({ctx.currency_sym})",
            ctx.result.sensitivity_irr_dlom,
            "IRR \\ DLOM", lambda v: f"{v:.0f}%", lambda v: f"{int(v)}%",
        )
        r += 2

    # ── WACC × Terminal Growth (DCF/SOTP) ──
    if ctx.result.sensitivity_dcf:
        n = sum(1 for x in [ctx.result.sensitivity_multiples, ctx.result.sensitivity_irr_dlom] if x)
        label_n = n + 1
        r = _write_sensitivity_table(
            ws, r,
            f"{'③' if label_n == 3 else '②' if label_n == 2 else '①'} WACC × 영구성장률 → 주당가치 ({ctx.unit})",
            ctx.result.sensitivity_dcf,
            "WACC \\ Tg", lambda v: f"{v:.1f}%", lambda v: f"{v:.1f}%",
            ref_value=ctx.result.weighted_value if method in ("sotp", "dcf_primary") else None,
        )
        r += 2

    # ── Primary method sensitivity ──
    if ctx.result.sensitivity_primary:
        row_fmt, col_fmt, corner = _sensitivity_format(method)
        n = sum(1 for x in [ctx.result.sensitivity_multiples, ctx.result.sensitivity_irr_dlom, ctx.result.sensitivity_dcf] if x)
        label_n = n + 1
        numbering = {1: "①", 2: "②", 3: "③", 4: "④"}.get(label_n, "")
        r = _write_sensitivity_table(
            ws, r,
            f"{numbering} {ctx.result.sensitivity_primary_label}",
            ctx.result.sensitivity_primary,
            corner, row_fmt, col_fmt,
        )
        r += 2

    # ── Reference value ──
    ref_label, ref_value = _get_ref_label_value(ctx)
    write_cell(ws, r, 1, f"참조: {ref_label} = {ref_value}",
               font=Font(italic=True, size=9, color="566573"))


def _sensitivity_format(method: str):
    """Return row/column format by valuation method for sensitivity tables."""
    if method == "ddm":
        return lambda v: f"{v:.1f}%", lambda v: f"{v:.1f}%", "Ke \\ Growth"
    elif method == "rim":
        return lambda v: f"{v:.1f}%", lambda v: f"{v:.1f}%", "Ke \\ Tg"
    elif method == "nav":
        return lambda v: f"{v:,.0f}", lambda v: f"{v:.0f}%", "재평가 \\ 할인율"
    elif method == "multiples":
        return lambda v: f"{v:.1f}x", lambda v: f"{v:.0f}%", "멀티플 \\ 할인율"
    elif method == "rnpv":
        return lambda v: f"{v:.1f}%", lambda v: f"{v:.1f}x", "DR \\ PoS Scale"
    return lambda v: f"{v}", lambda v: f"{v}", "Row \\ Col"


def _write_sensitivity_table(ws, r: int, title: str, data: list,
                              corner_label: str, row_fmt, col_fmt,
                              ref_value: int | None = None) -> int:
    """Write a generic 2D sensitivity table. Returns the last row number written."""
    lookup = {(x.row_val, x.col_val): x.value for x in data}
    row_range = sorted(set(x.row_val for x in data))
    col_range = sorted(set(x.col_val for x in data))

    if not row_range or not col_range:
        return r

    write_cell(ws, r, 1, title, font=SECTION_FONT)
    r += 1

    # Header
    write_cell(ws, r, 1, corner_label, fill=GRAY_FILL, font=HEADER_FONT)
    for j, col_v in enumerate(col_range, 2):
        write_cell(ws, r, j, col_fmt(col_v), fill=GRAY_FILL, font=HEADER_FONT)
        ws.column_dimensions[get_column_letter(j)].width = 12

    sens_start = r + 1
    for row_v in row_range:
        r += 1
        write_cell(ws, r, 1, row_fmt(row_v), fill=GRAY_FILL, font=HEADER_FONT)
        for j, col_v in enumerate(col_range, 2):
            val = lookup.get((row_v, col_v), 0)
            if ref_value is not None:
                fill = GREEN_FILL if val >= ref_value else RED_FILL
            else:
                fill = GREEN_FILL if val > 0 else RED_FILL
            write_cell(ws, r, j, val, fmt=NUM_FMT, fill=fill)
    sens_end = r

    # Heatmap
    if col_range and sens_end >= sens_start:
        end_col = get_column_letter(1 + len(col_range))
        ws.conditional_formatting.add(
            f"B{sens_start}:{end_col}{sens_end}", ColorScaleRule(
            start_type='min', start_color='FADBD8',
            mid_type='percentile', mid_value=50, mid_color='F5F6FA',
            end_type='max', end_color='D5F5E3',
        ))

    return r


def _get_ref_label_value(ctx: Ctx) -> tuple[str, str]:
    if ctx.method == "ddm" and ctx.result.ddm:
        return "DDM 주당가치", f"{ctx.result.ddm.equity_per_share:,}{ctx.currency_sym}"
    elif ctx.method == "rim" and ctx.result.rim:
        return "RIM 주당가치", f"{ctx.result.rim.per_share:,}{ctx.currency_sym}"
    elif ctx.method == "nav" and ctx.result.nav:
        return "NAV 주당가치", f"{ctx.result.nav.per_share:,}{ctx.currency_sym}"
    elif ctx.method == "multiples" and ctx.result.multiples_primary:
        return "Multiples 주당가치", f"{ctx.result.multiples_primary.per_share:,}{ctx.currency_sym}"
    elif ctx.result.dcf:
        return "DCF EV", f"{ctx.result.dcf.ev_dcf:,}{ctx.unit}"
    else:
        return "Total EV", f"{ctx.result.total_ev:,}{ctx.unit}"
