"""Sheet 5: Scenario Analysis — dynamic waterfall bridge."""

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ._ctx import Ctx
from ..excel_styles import (
    NAVY,
    BLUE_FILL,
    GREEN_FILL,
    RED_FILL,
    DRIVER_FILL,
    SECTION_FONT,
    TITLE_FONT,
    NOTE_FONT,
    NUM_FMT,
    style_header_row,
    write_cell,
)


def sheet_scenarios(ctx: Ctx):
    ws = ctx.wb.create_sheet("Scenario Analysis")
    ws.sheet_properties.tabColor = "F39C12"
    write_cell(ws, 1, 1, f"시나리오 분석 ({ctx.unit})", font=TITLE_FONT)

    sc_codes = ctx.sc_codes
    if not sc_codes:
        return

    r = 3
    sc_headers = ["항목"] + [f"{c}: {ctx.vi.scenarios[c].name}" for c in sc_codes]
    for c, h in enumerate(sc_headers, 1):
        write_cell(ws, r, c, h)
        ws.column_dimensions[get_column_letter(c)].width = 20 if c > 1 else 28
    style_header_row(ws, r, len(sc_headers))

    is_listed = ctx.vi.company.legal_status == "상장"
    any(ctx.vi.scenarios[c].dlom > 0 for c in sc_codes)

    # ── Scenario base assumptions ──
    r += 1
    write_cell(ws, r, 1, "확률")
    for i, sc_code in enumerate(sc_codes, 2):
        write_cell(ws, r, i, f"{ctx.vi.scenarios[sc_code].prob}%", fill=BLUE_FILL)

    # Method-specific key drivers (figures that differentiate scenarios)
    r = _write_scenario_drivers(ws, r, ctx)

    # DLOM -- only shown for private (unlisted) companies
    if not is_listed:
        r += 1
        write_cell(ws, r, 1, "DLOM")
        for i, sc_code in enumerate(sc_codes, 2):
            write_cell(ws, r, i, f"{ctx.vi.scenarios[sc_code].dlom}%", fill=BLUE_FILL)

    # IRR (private companies with CPS only)
    if any(ctx.vi.scenarios[c].irr is not None for c in sc_codes):
        r += 1
        write_cell(ws, r, 1, "FI IRR")
        for i, sc_code in enumerate(sc_codes, 2):
            irr = ctx.vi.scenarios[sc_code].irr
            write_cell(ws, r, i, f"{irr}%" if irr else "-", fill=BLUE_FILL)

    # Separator
    r += 1

    # ── Dynamic Equity Bridge (adjustments-based) ──
    r += 1
    write_cell(ws, r, 1, _ev_label(ctx.method), bold=True)
    for i, sc_code in enumerate(sc_codes, 2):
        sr = ctx.result.scenarios[sc_code]
        write_cell(ws, r, i, sr.total_ev, fmt=NUM_FMT)

    # Adjustments — Waterfall
    first_sr = ctx.result.scenarios[sc_codes[0]]
    for adj_idx, adj in enumerate(first_sr.adjustments):
        r += 1
        write_cell(ws, r, 1, f"(-) {adj.name}")
        for i, sc_code in enumerate(sc_codes, 2):
            sr = ctx.result.scenarios[sc_code]
            val = sr.adjustments[adj_idx].value if adj_idx < len(sr.adjustments) else 0
            write_cell(ws, r, i, val, fmt=NUM_FMT)

    # Equity Value
    r += 1
    write_cell(ws, r, 1, "Equity Value", bold=True)
    for i, sc_code in enumerate(sc_codes, 2):
        sr = ctx.result.scenarios[sc_code]
        fill = GREEN_FILL if sr.equity_value > 0 else RED_FILL
        write_cell(ws, r, i, sr.equity_value, fmt=NUM_FMT, bold=True, fill=fill)

    r += 1  # Separator

    # ── Per-share value ──
    r += 1
    write_cell(ws, r, 1, "적용 주식수")
    for i, sc_code in enumerate(sc_codes, 2):
        write_cell(ws, r, i, ctx.result.scenarios[sc_code].shares, fmt=NUM_FMT)

    if not is_listed:
        # Private company: show both pre/post DLOM
        r += 1
        write_cell(ws, r, 1, "주당 가치 (DLOM 전)")
        for i, sc_code in enumerate(sc_codes, 2):
            write_cell(ws, r, i, ctx.result.scenarios[sc_code].pre_dlom, fmt=NUM_FMT)

        r += 1
        write_cell(ws, r, 1, "주당 가치 (DLOM 후)", bold=True)
        for i, sc_code in enumerate(sc_codes, 2):
            sr = ctx.result.scenarios[sc_code]
            write_cell(
                ws,
                r,
                i,
                sr.post_dlom,
                fmt=NUM_FMT,
                bold=True,
                fill=GREEN_FILL if sr.post_dlom > 0 else None,
            )
    else:
        # Listed company: show single per-share value row
        r += 1
        write_cell(ws, r, 1, "주당 가치", bold=True)
        for i, sc_code in enumerate(sc_codes, 2):
            sr = ctx.result.scenarios[sc_code]
            write_cell(
                ws,
                r,
                i,
                sr.post_dlom,
                fmt=NUM_FMT,
                bold=True,
                fill=GREEN_FILL if sr.post_dlom > 0 else None,
            )

    r += 1
    write_cell(ws, r, 1, "확률가중 기여")
    for i, sc_code in enumerate(sc_codes, 2):
        write_cell(ws, r, i, ctx.result.scenarios[sc_code].weighted, fmt=NUM_FMT)

    # Probability-weighted conclusion
    r += 2
    write_cell(
        ws, r, 1, "확률가중 주당 가치", font=Font(bold=True, size=13, color=NAVY)
    )
    write_cell(
        ws,
        r,
        2,
        ctx.result.weighted_value,
        fmt=NUM_FMT,
        font=Font(bold=True, size=13, color="27AE60"),
        fill=GREEN_FILL,
    )
    write_cell(ws, r, 3, ctx.currency_sym, font=Font(bold=True, size=13, color=NAVY))

    # ── Scenario descriptions and probability rationale ──
    r += 3
    write_cell(ws, r, 1, "시나리오 설명 및 확률 배분 근거", font=SECTION_FONT)
    r += 1

    for sc_code in sc_codes:
        sc = ctx.vi.scenarios[sc_code]
        r += 1
        write_cell(ws, r, 1, f"{sc_code}: {sc.name}", font=Font(bold=True, color=NAVY))
        write_cell(ws, r, 2, f"확률 {sc.prob}%", fill=BLUE_FILL)
        if sc.desc:
            r += 1
            write_cell(ws, r, 1, f"  설명: {sc.desc}", font=NOTE_FONT)
            ws.column_dimensions[get_column_letter(1)].width = max(
                ws.column_dimensions[get_column_letter(1)].width or 0, 50
            )
        if sc.probability_rationale:
            r += 1
            write_cell(
                ws, r, 1, f"  확률 근거: {sc.probability_rationale}", font=NOTE_FONT
            )

    # ── News-to-driver mapping (AI analysis rationale) ──
    has_rationale = any(ctx.vi.scenarios[c].driver_rationale for c in sc_codes)
    if has_rationale:
        r += 3
        write_cell(ws, r, 1, "뉴스 → 드라이버 매핑 (AI 분석 근거)", font=SECTION_FONT)
        r += 1
        driver_labels = {
            "growth_adj_pct": "EBITDA 성장률 조정",
            "terminal_growth_adj": "영구성장률 조정",
            "wacc_adj": "WACC 조정",
            "market_sentiment_pct": "시장 심리 조정",
            "ddm_growth": "배당성장률",
            "ev_multiple": "적용 멀티플",
            "rim_roe_adj": "ROE 조정",
            "nav_discount": "지주할인율",
        }
        for sc_code in sc_codes:
            sc = ctx.vi.scenarios[sc_code]
            if not sc.driver_rationale:
                continue
            r += 1
            write_cell(
                ws, r, 1, f"{sc_code}: {sc.name}", font=Font(bold=True, color=NAVY)
            )
            for driver_name, rationale in sc.driver_rationale.items():
                r += 1
                label = driver_labels.get(driver_name, driver_name)
                write_cell(ws, r, 1, f"  {label}", font=NOTE_FONT)
                write_cell(ws, r, 2, rationale, font=NOTE_FONT)


def _write_scenario_drivers(ws, r: int, ctx: Ctx) -> int:
    """Write method-specific scenario key drivers and calculation rows."""
    sc_codes = ctx.sc_codes
    method = ctx.method
    calc_font = Font(italic=True, size=9, color="566573")

    if method == "multiples":
        mp = ctx.result.multiples_primary
        base_metric = mp.metric_value if mp else 0
        base_method = mp.primary_multiple_method if mp else "EV/EBITDA"

        # Driver: applied multiple
        r += 1
        write_cell(ws, r, 1, f"적용 {base_method}", bold=True)
        sc_multiples = []
        for i, sc_code in enumerate(sc_codes, 2):
            sc = ctx.vi.scenarios[sc_code]
            m = (
                sc.ev_multiple
                if sc.ev_multiple is not None
                else (mp.multiple if mp else 0)
            )
            sc_multiples.append(m)
            write_cell(ws, r, i, f"{m:.1f}x", fill=DRIVER_FILL, bold=True)

        # Calculation: base metric
        r += 1
        metric_label = {"EV/EBITDA": "EBITDA", "P/E": "순이익", "P/BV": "자기자본"}.get(
            base_method, "Metric"
        )
        write_cell(ws, r, 1, metric_label, font=calc_font)
        for i in range(len(sc_codes)):
            write_cell(ws, r, i + 2, base_metric, fmt=NUM_FMT, font=calc_font)

        # Calculation: formula = Metric x Multiple
        r += 1
        write_cell(ws, r, 1, f"  {metric_label} × {base_method}", font=calc_font)
        for i, m in enumerate(sc_multiples):
            write_cell(
                ws, r, i + 2, round(base_metric * m), fmt=NUM_FMT, font=calc_font
            )

    elif method == "ddm":
        ddm = ctx.result.ddm
        dps = ddm.dps if ddm else 0
        ke = ddm.ke if ddm else 0

        # Driver: dividend growth rate
        r += 1
        write_cell(ws, r, 1, "배당성장률 (g)", bold=True)
        sc_growths = []
        for i, sc_code in enumerate(sc_codes, 2):
            sc = ctx.vi.scenarios[sc_code]
            g = (
                sc.ddm_growth
                if sc.ddm_growth is not None
                else (ctx.vi.ddm_params.dividend_growth if ctx.vi.ddm_params else 0)
            )
            sc_growths.append(g)
            write_cell(ws, r, i, f"{g:.1f}%", fill=DRIVER_FILL, bold=True)

        # Calculation: Ke (reflecting per-scenario wacc_adj)
        r += 1
        write_cell(ws, r, 1, "Ke (자기자본비용)", font=calc_font)
        for i, sc_code in enumerate(sc_codes):
            sc_ke = ke + ctx.vi.scenarios[sc_code].wacc_adj
            write_cell(ws, r, i + 2, f"{sc_ke:.2f}%", font=calc_font)

        # Calculation: DPS
        r += 1
        write_cell(ws, r, 1, "DPS (주당배당금)", font=calc_font)
        for i in range(len(sc_codes)):
            write_cell(ws, r, i + 2, f"{dps:,.0f}원", font=calc_font)

        # Calculation: formula (reflecting per-scenario Ke)
        r += 1
        write_cell(ws, r, 1, "산식: DPS×(1+g) / (Ke-g)", font=calc_font)
        for i, g in enumerate(sc_growths):
            sc_ke = ke + ctx.vi.scenarios[sc_codes[i]].wacc_adj
            spread = sc_ke - g
            if spread > 0:
                val = round(dps * (1 + g / 100) / (spread / 100))
            else:
                val = 0
            write_cell(ws, r, i + 2, f"{val:,.0f}원", font=calc_font)

    elif method == "rim":
        ke = ctx.result.wacc.ke
        base_roes = ctx.vi.rim_params.roe_forecasts if ctx.vi.rim_params else []
        by = ctx.vi.base_year
        equity_bv = ctx.vi.consolidated[by].get("equity", 0)

        # Driver: ROE adjustment
        r += 1
        write_cell(ws, r, 1, "ROE 조정 (%p)", bold=True)
        sc_adjs = []
        for i, sc_code in enumerate(sc_codes, 2):
            sc = ctx.vi.scenarios[sc_code]
            adj = sc.rim_roe_adj
            sc_adjs.append(adj)
            label = f"{adj:+.1f}%p" if adj != 0 else "기본"
            write_cell(ws, r, i, label, fill=DRIVER_FILL, bold=True)

        # Calculation: applied ROE (year 1)
        if base_roes:
            r += 1
            write_cell(ws, r, 1, "적용 ROE (1년차)", font=calc_font)
            for i, adj in enumerate(sc_adjs):
                roe1 = base_roes[0] + adj
                write_cell(ws, r, i + 2, f"{roe1:.1f}%", font=calc_font)

        # Calculation: Ke (reflecting per-scenario wacc_adj)
        r += 1
        write_cell(ws, r, 1, "Ke (자기자본비용)", font=calc_font)
        for i, sc_code in enumerate(sc_codes):
            sc_ke = ke + ctx.vi.scenarios[sc_code].wacc_adj
            write_cell(ws, r, i + 2, f"{sc_ke:.2f}%", font=calc_font)

        # Calculation: RI Spread = ROE - Ke (per-scenario)
        if base_roes:
            r += 1
            write_cell(ws, r, 1, "RI Spread (ROE-Ke)", font=calc_font)
            for i, adj in enumerate(sc_adjs):
                sc_ke = ke + ctx.vi.scenarios[sc_codes[i]].wacc_adj
                spread = base_roes[0] + adj - sc_ke
                color = "27AE60" if spread > 0 else "E74C3C"
                write_cell(
                    ws,
                    r,
                    i + 2,
                    f"{spread:+.1f}%p",
                    font=Font(italic=True, size=9, color=color),
                )

        # Calculation: book value
        r += 1
        write_cell(ws, r, 1, "자기자본 (BV)", font=calc_font)
        for i in range(len(sc_codes)):
            write_cell(ws, r, i + 2, equity_bv, fmt=NUM_FMT, font=calc_font)

        # Calculation: formula description
        r += 1
        write_cell(ws, r, 1, "산식: BV + PV(RI) + PV(TV)", font=calc_font)

    elif method == "nav":
        nav_result = ctx.result.nav
        base_nav = nav_result.nav if nav_result else 0

        # Driver: holding company discount rate
        r += 1
        write_cell(ws, r, 1, "지주할인율", bold=True)
        sc_discounts = []
        for i, sc_code in enumerate(sc_codes, 2):
            sc = ctx.vi.scenarios[sc_code]
            sc_discounts.append(sc.nav_discount)
            write_cell(ws, r, i, f"{sc.nav_discount:.0f}%", fill=DRIVER_FILL, bold=True)

        # Calculation: NAV (pre-discount)
        r += 1
        write_cell(ws, r, 1, "NAV (할인 전)", font=calc_font)
        for i in range(len(sc_codes)):
            write_cell(ws, r, i + 2, base_nav, fmt=NUM_FMT, font=calc_font)

        # Calculation: formula = NAV x (1 - discount rate)
        r += 1
        write_cell(ws, r, 1, "산식: NAV × (1 - 할인율)", font=calc_font)
        for i, disc in enumerate(sc_discounts):
            discounted = round(base_nav * (1 - disc / 100))
            write_cell(ws, r, i + 2, discounted, fmt=NUM_FMT, font=calc_font)

    elif method == "rnpv":
        # rNPV: show growth_adj_pct (peak sales) and pos_override summary
        has_growth = any(ctx.vi.scenarios[c].growth_adj_pct != 0 for c in sc_codes)
        has_pos = any(ctx.vi.scenarios[c].pos_override for c in sc_codes)

        if has_growth:
            r += 1
            write_cell(ws, r, 1, "Peak Sales 조정", bold=True)
            for i, sc_code in enumerate(sc_codes, 2):
                adj = ctx.vi.scenarios[sc_code].growth_adj_pct
                label = f"{adj:+.0f}%" if adj != 0 else "기본"
                write_cell(ws, r, i, label, fill=DRIVER_FILL, bold=True)

        if has_pos:
            r += 1
            write_cell(ws, r, 1, "PoS Override", bold=True)
            for i, sc_code in enumerate(sc_codes, 2):
                pos_ov = ctx.vi.scenarios[sc_code].pos_override
                if pos_ov:
                    label = f"{len(pos_ov)}개 약물"
                else:
                    label = "기본"
                write_cell(ws, r, i, label, fill=DRIVER_FILL, bold=True)

            # Detail: list overridden drugs and PoS per scenario
            all_drug_names = set()
            for sc_code in sc_codes:
                pos_ov = ctx.vi.scenarios[sc_code].pos_override
                if pos_ov:
                    all_drug_names.update(pos_ov.keys())
            for drug_name in sorted(all_drug_names):
                r += 1
                short_name = drug_name.split("(")[0].strip()[:25]
                write_cell(ws, r, 1, f"  {short_name}", font=calc_font)
                for i, sc_code in enumerate(sc_codes, 2):
                    pos_ov = ctx.vi.scenarios[sc_code].pos_override
                    if pos_ov and drug_name in pos_ov:
                        write_cell(ws, r, i, f"{pos_ov[drug_name]:.0%}", font=calc_font)
                    else:
                        write_cell(ws, r, i, "-", font=calc_font)

        # WACC adjustment
        has_wacc = any(ctx.vi.scenarios[c].wacc_adj != 0 for c in sc_codes)
        if has_wacc:
            r += 1
            write_cell(ws, r, 1, "할인율 조정 (%p)", bold=True)
            for i, sc_code in enumerate(sc_codes, 2):
                adj = ctx.vi.scenarios[sc_code].wacc_adj
                label = f"{adj:+.2f}%p" if adj != 0 else "기본"
                write_cell(ws, r, i, label, fill=DRIVER_FILL, bold=True)

    elif method == "dcf_primary":
        has_growth = any(ctx.vi.scenarios[c].growth_adj_pct != 0 for c in sc_codes)
        has_tg = any(ctx.vi.scenarios[c].terminal_growth_adj != 0 for c in sc_codes)
        if has_growth:
            r += 1
            write_cell(ws, r, 1, "EBITDA 성장률 조정", bold=True)
            for i, sc_code in enumerate(sc_codes, 2):
                adj = ctx.vi.scenarios[sc_code].growth_adj_pct
                label = f"{adj:+.0f}%" if adj != 0 else "기본"
                write_cell(ws, r, i, label, fill=DRIVER_FILL, bold=True)
        if has_tg:
            r += 1
            write_cell(ws, r, 1, "영구성장률 조정", bold=True)
            for i, sc_code in enumerate(sc_codes, 2):
                adj = ctx.vi.scenarios[sc_code].terminal_growth_adj
                label = f"{adj:+.1f}%p" if adj != 0 else "기본"
                write_cell(ws, r, i, label, fill=DRIVER_FILL, bold=True)
        # Calculation: WACC (reflecting per-scenario wacc_adj)
        r += 1
        write_cell(ws, r, 1, "WACC", font=calc_font)
        for i, sc_code in enumerate(sc_codes):
            sc_wacc = ctx.result.wacc.wacc + ctx.vi.scenarios[sc_code].wacc_adj
            write_cell(ws, r, i + 2, f"{sc_wacc:.2f}%", font=calc_font)
        r += 1
        write_cell(ws, r, 1, "산식: DCF(FCFF, WACC, TGR)", font=calc_font)

    has_sentiment = any(ctx.vi.scenarios[c].market_sentiment_pct != 0 for c in sc_codes)
    if has_sentiment:
        r += 1
        write_cell(ws, r, 1, "시장 심리 조정", bold=True)
        for i, sc_code in enumerate(sc_codes, 2):
            adj = ctx.vi.scenarios[sc_code].market_sentiment_pct
            label = f"{adj:+.0f}%" if adj != 0 else "-"
            write_cell(ws, r, i, label, fill=DRIVER_FILL, bold=True)

    # WACC adjustment (cross-cutting -- also shown for methods other than DCF/DDM/RIM)
    if method in ("sotp", "multiples", "nav"):
        has_wacc_adj = any(ctx.vi.scenarios[c].wacc_adj != 0 for c in sc_codes)
        if has_wacc_adj:
            r += 1
            write_cell(ws, r, 1, "WACC 조정 (%p)", bold=True)
            for i, sc_code in enumerate(sc_codes, 2):
                adj = ctx.vi.scenarios[sc_code].wacc_adj
                label = f"{adj:+.2f}%p" if adj != 0 else "기본"
                write_cell(ws, r, i, label, fill=DRIVER_FILL, bold=True)

    return r


def _ev_label(method: str) -> str:
    """Return EV/Value label by valuation method."""
    labels = {
        "sotp": "SOTP EV",
        "dcf_primary": "DCF EV",
        "ddm": "DDM Equity Value",
        "rim": "RIM Equity Value",
        "nav": "NAV",
        "multiples": "Multiples EV",
        "rnpv": "rNPV Pipeline Value",
    }
    return labels.get(method, "Enterprise Value")
