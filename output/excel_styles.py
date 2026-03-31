"""Excel 스타일 상수 및 유틸리티 함수."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

# ── 색상 ──
NAVY = "1B2A4A"

# ── Fill ──
BLUE_FILL = PatternFill("solid", fgColor="E8EDF5")     # 입력 가정값
YELLOW_FILL = PatternFill("solid", fgColor="F0F2F8")    # 원본 데이터
GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")     # 양수 결과
RED_FILL = PatternFill("solid", fgColor="FDE8E8")       # 음수 결과
GRAY_FILL = PatternFill("solid", fgColor=NAVY)          # 헤더 (네이비)
DARK_FILL = PatternFill("solid", fgColor="2C3E50")      # 짙은 헤더
DRIVER_FILL = PatternFill("solid", fgColor="FFF3E0")    # 시나리오 드라이버 (연한 오렌지)

# ── Font ──
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
WHITE_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12, color=NAVY)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
NOTE_FONT = Font(size=9, color="566573")
RESULT_FONT = Font(bold=True, size=12, color="27AE60")

# ── Number Format ──
NUM_FMT = '#,##0'
PCT_FMT = '0.00%'
MULT_FMT = '0.0"x"'

# ── Border ──
THIN_BORDER = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC'),
)

BASE_BORDER = Border(
    left=Side(style='medium', color=NAVY),
    right=Side(style='medium', color=NAVY),
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='medium', color=NAVY),
)


def style_header_row(ws, row: int, max_col: int):
    """헤더 행 스타일 적용."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = GRAY_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def write_cell(ws, row: int, col: int, value, fmt=None, fill=None, font=None, bold=False):
    """셀 쓰기 + 스타일 적용."""
    cell = ws.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    elif bold:
        cell.font = Font(bold=True, color=NAVY)
    cell.border = THIN_BORDER
    return cell
