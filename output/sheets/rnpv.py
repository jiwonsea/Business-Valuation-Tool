"""Sheet: rNPV Pipeline Valuation — Pipeline Summary + Revenue Curves."""

from openpyxl.utils import get_column_letter

from ._ctx import Ctx
from ..excel_styles import (
    BLUE_FILL,
    YELLOW_FILL,
    GREEN_FILL,
    RED_FILL,
    GRAY_FILL,
    SECTION_FONT,
    TITLE_FONT,
    NOTE_FONT,
    RESULT_FONT,
    NUM_FMT,
    PCT_FMT,
    style_header_row,
    write_cell,
)


def valuation_rnpv(ctx: Ctx):
    """rNPV Pipeline Summary + Revenue Curves sheets."""
    rnpv = ctx.result.rnpv
    if not rnpv:
        ws = ctx.wb.create_sheet("rNPV Pipeline")
        write_cell(ws, 1, 1, "rNPV 결과 없음", font=SECTION_FONT)
        return

    _sheet_pipeline_summary(ctx, rnpv)
    _sheet_revenue_curves(ctx, rnpv)


def _sheet_pipeline_summary(ctx: Ctx, rnpv):
    """Pipeline Summary sheet: per-drug rNPV breakdown."""
    ws = ctx.wb.create_sheet("rNPV Pipeline")
    ws.sheet_properties.tabColor = "8E44AD"
    ws.column_dimensions["A"].width = 28

    write_cell(ws, 1, 1, f"rNPV 파이프라인 밸류에이션 ({ctx.unit})", font=TITLE_FONT)

    # Pipeline summary table
    r = 3
    write_cell(ws, r, 1, "파이프라인 약물별 rNPV", font=SECTION_FONT)
    r += 1

    headers = [
        "Drug",
        "Phase",
        "Indication",
        "Peak Sales",
        "PoS",
        "NPV",
        "rNPV",
        "비중",
    ]
    for c, h in enumerate(headers, 1):
        write_cell(ws, r, c, h)
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["C"].width = 22
    style_header_row(ws, r, len(headers))

    phase_labels = {
        "preclinical": "Preclinical",
        "phase1": "Phase 1",
        "phase2": "Phase 2",
        "phase3": "Phase 3",
        "filed": "Filed/NDA",
        "approved": "Approved",
    }

    for dr in rnpv.drug_results:
        r += 1
        rnpv_pct = dr.rnpv / rnpv.total_rnpv if rnpv.total_rnpv != 0 else 0
        is_approved = dr.phase == "approved"
        write_cell(ws, r, 1, dr.name, bold=is_approved)
        write_cell(ws, r, 2, phase_labels.get(dr.phase, dr.phase))
        write_cell(ws, r, 3, dr.indication)
        write_cell(ws, r, 4, dr.peak_sales, fmt=NUM_FMT)
        write_cell(
            ws,
            r,
            5,
            dr.success_prob,
            fmt=PCT_FMT,
            fill=GREEN_FILL
            if dr.success_prob >= 0.8
            else YELLOW_FILL
            if dr.success_prob >= 0.5
            else None,
        )
        write_cell(ws, r, 6, dr.npv_unadjusted, fmt=NUM_FMT)
        write_cell(
            ws, r, 7, dr.rnpv, fmt=NUM_FMT, fill=GREEN_FILL if dr.rnpv > 0 else RED_FILL
        )
        write_cell(ws, r, 8, rnpv_pct, fmt=PCT_FMT)

    # Totals row
    r += 1
    write_cell(ws, r, 1, "합계", bold=True)
    write_cell(
        ws, r, 4, sum(dr.peak_sales for dr in rnpv.drug_results), fmt=NUM_FMT, bold=True
    )
    write_cell(
        ws,
        r,
        6,
        sum(dr.npv_unadjusted for dr in rnpv.drug_results),
        fmt=NUM_FMT,
        bold=True,
    )
    write_cell(ws, r, 7, rnpv.total_rnpv, fmt=NUM_FMT, bold=True, fill=GREEN_FILL)
    write_cell(ws, r, 8, 1.0, fmt=PCT_FMT, bold=True)

    # Equity Bridge
    r += 2
    write_cell(ws, r, 1, "Equity Bridge", font=SECTION_FONT)
    r += 1

    bridge_items = [
        ("Total rNPV (약물 합계)", rnpv.total_rnpv),
        ("(-) PV(R&D Costs)", rnpv.r_and_d_cost_pv),
        ("Pipeline Value", rnpv.pipeline_value),
        ("(-) Net Debt", ctx.vi.net_debt),
        ("Equity Value", rnpv.pipeline_value - ctx.vi.net_debt),
    ]
    for label, val in bridge_items:
        is_total = label == "Equity Value"
        write_cell(ws, r, 1, label, bold=is_total)
        write_cell(
            ws,
            r,
            2,
            val,
            fmt=NUM_FMT,
            bold=is_total,
            fill=GREEN_FILL if is_total else YELLOW_FILL,
        )
        write_cell(ws, r, 3, ctx.unit)
        r += 1

    # Per-share
    r += 1
    write_cell(ws, r, 1, "주당 rNPV 가치", bold=True)
    write_cell(ws, r, 2, rnpv.per_share, fmt=NUM_FMT, font=RESULT_FONT, fill=GREEN_FILL)
    write_cell(ws, r, 3, ctx.currency_sym)

    # Key assumptions
    r += 2
    write_cell(ws, r, 1, "주요 가정", font=SECTION_FONT)
    r += 1
    assumptions = [
        ("할인율 (WACC)", f"{rnpv.discount_rate:.2f}%"),
        (
            "영업이익률",
            f"{ctx.vi.rnpv_params.default_margin:.0%}" if ctx.vi.rnpv_params else "-",
        ),
        (
            "법인세율",
            f"{ctx.vi.rnpv_params.tax_rate:.0%}" if ctx.vi.rnpv_params else "-",
        ),
        (
            "특허만료 후 감소율",
            f"{ctx.vi.rnpv_params.decline_rate:.0f}%/yr" if ctx.vi.rnpv_params else "-",
        ),
        (
            "R&D 비용 (별도공제)",
            f"{ctx.vi.rnpv_params.r_and_d_cost:,} {ctx.unit}/yr"
            if ctx.vi.rnpv_params
            else "-",
        ),
    ]
    for label, val in assumptions:
        write_cell(ws, r, 1, label, fill=BLUE_FILL)
        write_cell(ws, r, 2, val, fill=BLUE_FILL)
        r += 1

    # Reverse rNPV section
    rv = ctx.result.reverse_rnpv
    if rv:
        r += 2
        write_cell(ws, r, 1, "역방향 rNPV (시장 내재 가정)", font=SECTION_FONT)
        r += 1
        write_cell(ws, r, 1, "모델 EV", fill=GRAY_FILL)
        write_cell(ws, r, 2, rv.model_ev, fmt=NUM_FMT, fill=GRAY_FILL)
        write_cell(ws, r, 3, ctx.unit, fill=GRAY_FILL)
        r += 1
        write_cell(ws, r, 1, "시장 EV (Target)", fill=GRAY_FILL)
        write_cell(ws, r, 2, rv.target_ev, fmt=NUM_FMT, fill=GRAY_FILL)
        write_cell(ws, r, 3, ctx.unit, fill=GRAY_FILL)
        r += 1
        write_cell(ws, r, 1, "괴리율")
        write_cell(
            ws,
            r,
            2,
            rv.gap_pct / 100,
            fmt=PCT_FMT,
            fill=RED_FILL if abs(rv.gap_pct) >= 20 else YELLOW_FILL,
        )

        if rv.implied_pos_scale is not None:
            r += 2
            write_cell(ws, r, 1, "시장 내재 PoS 배수", bold=True)
            write_cell(ws, r, 2, f"{rv.implied_pos_scale:.3f}x")
            r += 1
            headers = ["Drug", "Base PoS", "Implied PoS", "Δ"]
            for c, h in enumerate(headers, 1):
                write_cell(ws, r, c, h)
            style_header_row(ws, r, len(headers))
            for d in rv.implied_pos_per_drug:
                r += 1
                write_cell(ws, r, 1, d.name)
                write_cell(ws, r, 2, d.base_value, fmt=PCT_FMT)
                write_cell(
                    ws,
                    r,
                    3,
                    d.implied_value,
                    fmt=PCT_FMT,
                    fill=GREEN_FILL if d.implied_value > d.base_value else RED_FILL,
                )
                delta = d.implied_value - d.base_value
                write_cell(ws, r, 4, delta, fmt=PCT_FMT)

        if rv.implied_peak_scale is not None:
            r += 2
            write_cell(ws, r, 1, "시장 내재 Peak Sales 배수", bold=True)
            write_cell(ws, r, 2, f"{rv.implied_peak_scale:.3f}x")
            r += 1
            headers = ["Drug", "Base Peak", "Implied Peak", "Δ%"]
            for c, h in enumerate(headers, 1):
                write_cell(ws, r, c, h)
            style_header_row(ws, r, len(headers))
            for d in rv.implied_peak_per_drug:
                r += 1
                write_cell(ws, r, 1, d.name)
                write_cell(ws, r, 2, int(d.base_value), fmt=NUM_FMT)
                write_cell(
                    ws,
                    r,
                    3,
                    int(d.implied_value),
                    fmt=NUM_FMT,
                    fill=GREEN_FILL if d.implied_value > d.base_value else RED_FILL,
                )
                delta_pct = (
                    (d.implied_value - d.base_value) / d.base_value
                    if d.base_value
                    else 0
                )
                write_cell(ws, r, 4, delta_pct, fmt=PCT_FMT)

        if rv.implied_discount_rate is not None:
            r += 2
            dr_current = rnpv.discount_rate
            write_cell(ws, r, 1, "시장 내재 할인율", bold=True)
            write_cell(
                ws, r, 2, f"{rv.implied_discount_rate:.2f}% (현재 {dr_current:.2f}%)"
            )

        # Per-drug independent PoS (solo analysis)
        solo_drugs = [d for d in rv.implied_pos_solo if not d.skipped]
        if solo_drugs:
            r += 2
            direction = "시장 낙관" if rv.gap_pct < 0 else "시장 비관"
            write_cell(
                ws, r, 1, f"약물별 독립 PoS 분석 ({direction})", font=SECTION_FONT
            )
            r += 1
            headers = [
                "Drug",
                "Phase",
                "Base PoS",
                "Implied PoS",
                "Solvable",
                "Max Contribution",
            ]
            for c, h in enumerate(headers, 1):
                write_cell(ws, r, c, h)
            style_header_row(ws, r, len(headers))
            for d in solo_drugs:
                r += 1
                write_cell(ws, r, 1, d.name)
                write_cell(ws, r, 2, d.phase)
                write_cell(ws, r, 3, d.base_pos, fmt=PCT_FMT)
                if d.solvable and d.implied_pos is not None:
                    write_cell(
                        ws,
                        r,
                        4,
                        d.implied_pos,
                        fmt=PCT_FMT,
                        fill=GREEN_FILL if d.implied_pos > d.base_pos else RED_FILL,
                    )
                else:
                    write_cell(ws, r, 4, "N/A")
                write_cell(ws, r, 5, "Y" if d.solvable else "N")
                write_cell(ws, r, 6, d.max_ev_contribution, fmt=NUM_FMT)
            r += 1
            write_cell(
                ws,
                r,
                1,
                "※ 각 약물은 다른 약물을 현재 가정으로 고정한 독립 분석. 결과는 합산 불가.",
                font=NOTE_FONT,
            )

    # Tornado chart data (per-drug peak sales ±20% impact)
    tornado = ctx.result.rnpv_tornado
    if tornado:
        r += 2
        write_cell(ws, r, 1, "Tornado 분석 (Peak Sales ±20% 영향)", font=SECTION_FONT)
        r += 1
        headers = ["Drug", "Low (-20%)", "Base", "High (+20%)", "Swing"]
        for c, h in enumerate(headers, 1):
            write_cell(ws, r, c, h)
            ws.column_dimensions[get_column_letter(c)].width = 16
        ws.column_dimensions["A"].width = 28
        style_header_row(ws, r, len(headers))

        for t in tornado:
            r += 1
            swing = t.high_value - t.low_value
            write_cell(ws, r, 1, t.name)
            write_cell(ws, r, 2, t.low_value, fmt=NUM_FMT, fill=RED_FILL)
            write_cell(ws, r, 3, t.base_value, fmt=NUM_FMT)
            write_cell(ws, r, 4, t.high_value, fmt=NUM_FMT, fill=GREEN_FILL)
            write_cell(
                ws, r, 5, swing, fmt=NUM_FMT, fill=YELLOW_FILL if swing > 0 else None
            )


def _sheet_revenue_curves(ctx: Ctx, rnpv):
    """Revenue Curves sheet: year-by-year revenue projection per drug (chart-ready data)."""
    ws = ctx.wb.create_sheet("Revenue Curves")
    ws.sheet_properties.tabColor = "2E86C1"
    ws.column_dimensions["A"].width = 8

    write_cell(ws, 1, 1, f"파이프라인 연도별 매출 추정 ({ctx.unit})", font=TITLE_FONT)

    # Find max curve length
    drugs_with_curves = [dr for dr in rnpv.drug_results if dr.revenue_curve]
    if not drugs_with_curves:
        write_cell(ws, 3, 1, "매출 추정 데이터 없음", font=NOTE_FONT)
        return

    max_years = max(len(dr.revenue_curve) for dr in drugs_with_curves)
    base_year = ctx.by

    # Header row: Year | Drug1 | Drug2 | ... | Total
    r = 3
    write_cell(ws, r, 1, "Year")
    for c, dr in enumerate(drugs_with_curves, 2):
        short_name = dr.name.split("(")[0].strip()[:20]
        write_cell(ws, r, c, short_name)
        ws.column_dimensions[get_column_letter(c)].width = 14
    total_col = len(drugs_with_curves) + 2
    write_cell(ws, r, total_col, "Total")
    ws.column_dimensions[get_column_letter(total_col)].width = 14
    style_header_row(ws, r, total_col)

    # Data rows
    for yr in range(max_years):
        r += 1
        write_cell(ws, r, 1, base_year + yr)
        row_total = 0
        for c, dr in enumerate(drugs_with_curves, 2):
            rev = dr.revenue_curve[yr] if yr < len(dr.revenue_curve) else 0
            write_cell(ws, r, c, rev, fmt=NUM_FMT)
            row_total += rev
        write_cell(ws, r, total_col, row_total, fmt=NUM_FMT, bold=True)

    # Peak revenue summary below
    r += 2
    write_cell(ws, r, 1, "Peak Revenue 요약", font=SECTION_FONT)
    r += 1
    headers = ["Drug", "Peak Sales", "PoS", "Risk-adj Peak"]
    for c, h in enumerate(headers, 1):
        write_cell(ws, r, c, h)
    style_header_row(ws, r, len(headers))

    for dr in rnpv.drug_results:
        r += 1
        write_cell(ws, r, 1, dr.name.split("(")[0].strip()[:25])
        write_cell(ws, r, 2, dr.peak_sales, fmt=NUM_FMT)
        write_cell(ws, r, 3, dr.success_prob, fmt=PCT_FMT)
        write_cell(ws, r, 4, round(dr.peak_sales * dr.success_prob), fmt=NUM_FMT)
